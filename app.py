#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
雨课堂考勤分析 - Streamlit 在线工具
"""

import sys
import os
import base64
from collections import OrderedDict
from io import BytesIO

import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="雨课堂考勤分析", layout="wide")

# ── 核心解析函数（与本地版一致） ──


def col_idx(col):
    i = 0
    for ch in col:
        i = i * 26 + (ord(ch) - 64)
    return i - 1


def parse_summary(wb):
    if not wb.sheetnames:
        raise ValueError('Excel 文件中没有找到任何工作表。')

    ws = wb[wb.sheetnames[0]]

    row2 = {}
    for cell in ws[2]:
        if cell.value:
            row2[cell.column_letter] = str(cell.value)

    if not row2:
        raise ValueError('汇总表第 2 行为空，未找到子表头信息。请确认上传的是雨课堂批量导出文件（含汇总页）。')

    row1 = {}
    for cell in ws[1]:
        if cell.value:
            row1[cell.column_letter] = str(cell.value)

    if not row1:
        raise ValueError('汇总表第 1 行为空，未找到课堂名称合并表头。')

    sign_cols = sorted(
        [c for c, v in row2.items() if v == '签到方式'],
        key=col_idx
    )
    if not sign_cols:
        raise ValueError('未找到"签到方式"列。请确认上传的是雨课堂导出的"汇总-数据表"文件。')

    score_cols = sorted(
        [c for c, v in row2.items() if v.startswith('得分')],
        key=col_idx
    )
    session_headers = sorted(row1.keys(), key=col_idx)

    def map_cols_to_sessions(cols):
        result = OrderedDict()
        keys = []
        for c in cols:
            ci = col_idx(c)
            nearest = None
            nearest_i = -1
            for sh in session_headers:
                hi = col_idx(sh)
                if hi <= ci and hi > nearest_i:
                    nearest = row1[sh]
                    nearest_i = hi
            if nearest:
                result[nearest] = c
                keys.append(nearest)
        return result, keys

    session_sign_map, session_keys = map_cols_to_sessions(sign_cols)
    session_score_map, _ = map_cols_to_sessions(score_cols)

    students = []
    for row in ws.iter_rows(min_row=3, values_only=False):
        cells = {}
        for c in row:
            if c.value is None:
                continue
            try:
                cells[c.column_letter] = str(c.value)
            except AttributeError:
                pass
        sid = cells.get('A', '')
        name = cells.get('D', '')
        if not sid and not name:
            continue

        rec = {
            'id': sid,
            'name': name,
            'dept': cells.get('B', ''),
            'cls': cells.get('C', ''),
            'attendance': {},
            'scores': {},
        }
        for sk, sc in session_sign_map.items():
            rec['attendance'][sk] = cells.get(sc, '')
        for sk, sc in session_score_map.items():
            raw = cells.get(sc, '')
            try:
                rec['scores'][sk] = int(float(raw)) if raw else None
            except ValueError:
                rec['scores'][sk] = None
        students.append(rec)

    return session_keys, session_sign_map, session_score_map, students


def parse_sub_sheets(wb, session_keys):
    leave_data = {}
    idx = 0
    found = False

    for name in wb.sheetnames:
        if '课堂情况' not in name:
            continue
        found = True
        if idx >= len(session_keys):
            break

        session_key = session_keys[idx]
        idx += 1
        ws = wb[name]

        for row in ws.iter_rows(min_row=4, values_only=False):
            cells = {}
            for c in row:
                if c.value is None:
                    continue
                try:
                    cells[c.column_letter] = str(c.value)
                except AttributeError:
                    pass
            sid = cells.get('A', '')
            remark = cells.get('G', '')
            if sid and remark in ('病假', '事假'):
                leave_data[(session_key, sid)] = remark

    if not found:
        raise ValueError('未找到名称包含"课堂情况"的子表。请确认上传的是雨课堂批量导出文件（含每次课的课堂情况子表）。')

    return leave_data


def parse_single_session(wb):
    """Parse a single-session export file (no summary page, just one 课堂情况 sheet)."""
    ws = wb[wb.sheetnames[0]]

    raw_name = str(ws.cell(1, 1).value or '').strip()
    if not raw_name:
        raise ValueError('文件第 1 行为空，未找到课堂名称。')

    row3_col5 = str(ws.cell(3, 5).value or '')
    if '签到方式' not in row3_col5:
        raise ValueError('第 3 行未找到"签到方式"列，请确认是雨课堂导出的课堂数据表。')

    students = []
    for row_idx in range(5, ws.max_row + 1):
        sid = str(ws.cell(row_idx, 1).value or '').strip()
        name = str(ws.cell(row_idx, 4).value or '').strip()
        if not sid and not name:
            continue
        students.append({
            'id': sid,
            'name': name,
            'dept': '',
            'cls': '',
            'attendance': {},
            'scores': {},
        })
        # Directly read sign and remark from this row
        students[-1]['_sign'] = str(ws.cell(row_idx, 5).value or '')
        students[-1]['_remark'] = str(ws.cell(row_idx, 7).value or '')

    session_keys = [raw_name]
    session_sign_map = OrderedDict([(raw_name, raw_name)])
    session_score_map = OrderedDict()

    leave_data = {}
    for s in students:
        s['attendance'][raw_name] = s.pop('_sign')
        remark = s.pop('_remark')
        if remark in ('病假', '事假'):
            leave_data[(raw_name, s['id'])] = remark

    return session_keys, session_sign_map, session_score_map, students, leave_data


def parse_file(wb):
    """Auto-detect file type (full export or single session) and parse."""
    # Detect: if first sheet name contains 课堂情况 and there's only 1 sheet, it's single-session
    if len(wb.sheetnames) == 1 and '课堂情况' in wb.sheetnames[0]:
        return parse_single_session(wb)
    # Otherwise treat as full export
    sk, ssm, sscm, students = parse_summary(wb)
    ld = parse_sub_sheets(wb, sk)
    return sk, ssm, sscm, students, ld


def generate_output(session_keys, session_sign_map, session_score_map, students, leave_data):
    """Generate output Excel bytes + summary data from parsed data."""

    present_set = {'扫二维码', '“正在上课”提示', '教师添加', '课堂暗号'}
    session_count = len(session_sign_map)

    # ── Build data for Sheet 1 ──
    out_rows = []
    for s in students:
        row = [s['id'], s['name']]
        absent_cnt = 0
        excused_cnt = 0
        for sk in session_sign_map:
            raw = s['attendance'].get(sk, '')
            if raw in present_set:
                status = '上课'
            elif raw == '未上课':
                leave = leave_data.get((sk, s['id']))
                if leave == '病假':
                    status = '病假'
                    excused_cnt += 1
                elif leave == '事假':
                    status = '事假'
                    excused_cnt += 1
                else:
                    status = '旷课'
                    absent_cnt += 1
            else:
                status = raw
            row.append(status)
        row.append(f'{absent_cnt / session_count:.0%}')
        row.append(f'{(absent_cnt + excused_cnt) / session_count:.0%}')
        out_rows.append(row)

    # ── Write output workbook ──
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    color_map = {'上课': green_fill, '病假': yellow_fill, '事假': yellow_fill, '旷课': red_fill}
    gold_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')

    owb = Workbook()
    ows = owb.active
    ows.title = '考勤明细'
    headers = ['学号', '姓名'] + list(session_sign_map.keys()) + ['无故旷课率', '总旷课率']
    ows.append(headers)
    for row in out_rows:
        ows.append(row)
    for row in ows.iter_rows(min_row=2, max_row=ows.max_row,
                             min_col=3, max_col=2 + session_count):
        for cell in row:
            fill = color_map.get(cell.value)
            if fill:
                cell.fill = fill

    # ── Sheet 2: 课堂表现 ──
    ows2 = owb.create_sheet('课堂表现')
    score_headers = ['学号', '姓名'] + [k for k in session_score_map.keys()] + ['总分']
    ows2.append(score_headers)
    totals = []
    for s in students:
        row = [s['id'], s['name']]
        total = 0
        for sk in session_score_map:
            v = s['scores'].get(sk)
            if v is not None:
                row.append(v)
                total += v
            else:
                row.append('')
        row.append(total)
        totals.append(total)
        ows2.append(row)

    # Top 10% gold highlight
    sorted_totals = sorted(totals, reverse=True)
    n_top = max(1, round(len(sorted_totals) * 0.1))
    threshold = sorted_totals[n_top - 1] if n_top <= len(sorted_totals) else sorted_totals[-1]
    for row in ows2.iter_rows(min_row=2, max_row=ows2.max_row):
        if row[-1].value is not None and row[-1].value >= threshold:
            for cell in row:
                cell.fill = gold_fill

    buf = BytesIO()
    owb.save(buf)
    buf.seek(0)

    # ── Summary stats ──
    summary_lines = []
    for sk in session_sign_map:
        attend = absent = sick = personal = 0
        for s in students:
            raw = s['attendance'].get(sk, '')
            if raw in present_set:
                attend += 1
            elif raw == '未上课':
                leave = leave_data.get((sk, s['id']))
                if leave == '病假':
                    sick += 1
                elif leave == '事假':
                    personal += 1
                else:
                    absent += 1
        summary_lines.append({
            'session': sk,
            '上课': attend,
            '旷课': absent,
            '病假': sick,
            '事假': personal,
        })

    total_absent = sum(1 for s in students for sk in session_sign_map
                       if s['attendance'].get(sk) == '未上课'
                       and (sk, s['id']) not in leave_data)
    total_sick = sum(1 for v in leave_data.values() if v == '病假')
    total_personal = sum(1 for v in leave_data.values() if v == '事假')

    # ── Build processed data for preview ──
    session_names = list(session_sign_map.keys())
    preview_attendance = []
    for s in students:
        row = {'学号': s['id'], '姓名': s['name']}
        absent_cnt = 0
        excused_cnt = 0
        for sk in session_names:
            raw = s['attendance'].get(sk, '')
            if raw in present_set:
                status = '上课'
            elif raw == '未上课':
                leave = leave_data.get((sk, s['id']))
                if leave == '病假':
                    status = '病假'
                    excused_cnt += 1
                elif leave == '事假':
                    status = '事假'
                    excused_cnt += 1
                else:
                    status = '旷课'
                    absent_cnt += 1
            else:
                status = raw
            row[sk] = status
        row['无故旷课率'] = f'{absent_cnt / session_count:.0%}'
        row['总旷课率'] = f'{(absent_cnt + excused_cnt) / session_count:.0%}'
        preview_attendance.append(row)

    score_names = list(session_score_map.keys())
    preview_scores = []
    for s in students:
        row = {'学号': s['id'], '姓名': s['name']}
        total = 0
        for k in score_names:
            v = s['scores'].get(k)
            if v is not None:
                row[k] = v
                total += v
            else:
                row[k] = ''
        row['总分'] = total
        preview_scores.append(row)

    return buf, {
        'students': students,
        'session_count': session_count,
        'total_absent': total_absent,
        'total_sick': total_sick,
        'total_personal': total_personal,
        'summary_lines': summary_lines,
        'preview_headers': ['学号', '姓名'] + session_names + ['无故旷课率', '总旷课率'],
        'preview_attendance': preview_attendance,
        'score_headers': ['学号', '姓名'] + score_names + ['总分'],
        'preview_scores': preview_scores,
        'session_keys': session_keys,
        'leave_data': leave_data,
        'process_score_buf': generate_process_score_sheet(students, session_keys, leave_data),
    }


# ── 过程性成绩记载表 ──

PRESENT_SET = {'扫二维码', '“正在上课”提示', '教师添加', '课堂暗号'}


def generate_process_score_sheet(students, session_keys, leave_data):
    """生成过程性成绩记载表（✓/✗/△），与样表格式一致"""
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    owb = Workbook()
    ows = owb.active
    ows.title = '过程性成绩记载表'

    # 列布局：序号(A) 学号(B) 姓名(C) 性别(D) 专业(E) 班级(F) 课堂1(G) ... 成绩 备注
    n_sessions = len(session_keys)
    session_start = 7  # column G
    score_col = max(session_start + n_sessions, 26)  # at least Z
    remark_col = score_col + 1

    headers = ['序号', '学号', '姓名', '性别', '专业', '班级'] + list(session_keys)
    while len(headers) < score_col - 1:
        headers.append('')
    headers.append('成绩')
    headers.append('备注')
    ows.append(headers)

    for i, s in enumerate(students, start=1):
        row = [i, s['id'], s['name'], '', '', s['cls']]
        for sk in session_keys:
            raw = s['attendance'].get(sk, '')
            if raw in PRESENT_SET:
                row.append('✓')
            elif raw == '未上课':
                leave = leave_data.get((sk, s['id']))
                row.append('△' if leave in ('病假', '事假') else '✗')
            else:
                row.append(raw)
        while len(row) < score_col - 1:
            row.append('')
        row.append('')  # 成绩
        row.append('')  # 备注
        ows.append(row)

    # ── 样式：宋体 9pt、居中、细边框 ──
    font_song = Font(name='SimSun', size=9)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for row in ows.iter_rows(min_row=1, max_row=ows.max_row,
                              min_col=1, max_col=remark_col):
        for cell in row:
            cell.font = font_song
            cell.alignment = align_center
            cell.border = thin_border

    # ── 列宽（参照样表）──
    col_widths = {
        'A': 2.85, 'B': 12.75, 'C': 11.26, 'D': 3.25,
        'E': 15.87, 'F': 15.47,
    }
    for col_letter, width in col_widths.items():
        ows.column_dimensions[col_letter].width = width
    for i in range(session_start, score_col):
        ows.column_dimensions[get_column_letter(i)].width = 3.69
    ows.column_dimensions[get_column_letter(score_col)].width = 5.38
    ows.column_dimensions[get_column_letter(remark_col)].width = 9.77

    # ── 行高 ──
    ows.row_dimensions[1].height = 22.6
    for r in range(2, ows.max_row + 1):
        ows.row_dimensions[r].height = 14.3

    buf = BytesIO()
    owb.save(buf)
    buf.seek(0)
    return buf


# ── Streamlit UI ──

DEMO_FILE = '示例表格.xlsx'


def _sample_link():
    """Return base64 data URI for the sample file download link."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), DEMO_FILE)
    if not os.path.exists(path):
        return ''
    with open(path, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode()
    return f'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}'


st.markdown("""
<style>
    .block-container { padding-top: 2.5rem; padding-bottom: 3rem; }
</style>
""", unsafe_allow_html=True)

st.title("长江雨课堂考勤数据分析工具")

st.markdown(f"""
上传雨课堂导出的 Excel 文件，自动生成考勤明细和课堂表现统计。

**适用平台：** [长江雨课堂](https://changjiang.yuketang.cn/web/?index)

**支持的文件类型：**
- 批量导出的汇总数据表（含汇总页 + 课堂情况子表）
- 单次课导出的课堂数据表（自动识别）

**原始文件获取方式：**
1. 登录长江雨课堂官网，进入你授课的班级
2. 点击 **批量导出数据**
3. 筛选需要统计的课堂记录
4. 下载导出的 Excel 文件（即 `.xlsx` 格式的汇总数据表）。<a href="{_sample_link()}" download="{DEMO_FILE}">下载示例表格</a>

**上传后：**
- 自动解析并分 tab 展示考勤明细、课堂表现得分和统计摘要
- 支持单文件分析和两表合并（如理论班 + 实验班）
- 提供两种下载：考勤明细 Excel（颜色标注）和过程性成绩记载表（✓/✗/△）
""", unsafe_allow_html=True)

st.markdown("""
<style>
    button[kind='segmented_control'], button[kind='segmented_controlActive'] {
        font-size: 1.8rem !important;
        padding: 1rem 2.5rem !important;
    }
</style>
""", unsafe_allow_html=True)

mode = st.segmented_control("模式", ["单文件模式", "合并模式"],
                            default="单文件模式", label_visibility="collapsed")


def show_results(buf, info):
    """显示预览表格和下载按钮"""
    st.subheader(f"考勤概况 — 共 {info['session_count']} 次课，{len(info['students'])} 名学生")

    tab1, tab2, tab3 = st.tabs(["考勤明细", "课堂表现", "统计摘要"])

    with tab1:
        st.info("完整数据请下载 Excel 文件查看。下面预览前 10 名学生：")
        df = pd.DataFrame(info['preview_attendance'][:10])
        styled = df.style.map(
            lambda v: 'background-color: #C6EFCE' if v == '上课'
            else 'background-color: #FFEB9C' if v in ('病假', '事假')
            else 'background-color: #FFC7CE' if v == '旷课'
            else ''
        )
        st.dataframe(styled, use_container_width=True)

    with tab2:
        st.info("课堂表现得分详情请下载 Excel 文件查看。下面预览前 10 名学生：")
        all_scores = info['preview_scores']
        totals = [s['总分'] for s in all_scores]
        sorted_totals = sorted(totals, reverse=True)
        n_top = max(1, round(len(sorted_totals) * 0.1))
        threshold = sorted_totals[n_top - 1]

        df_scores = pd.DataFrame(all_scores[:10])
        styled_scores = df_scores.style.apply(
            lambda row: ['background-color: #FFD700'] * len(row)
            if row['总分'] >= threshold else [''] * len(row),
            axis=1,
        )
        st.dataframe(styled_scores, use_container_width=True)

    with tab3:
        col1, col2, col3 = st.columns(3)
        col1.metric("旷课总人次", info['total_absent'])
        col2.metric("病假总人次", info['total_sick'])
        col3.metric("事假总人次", info['total_personal'])

        st.subheader("每次课统计")
        summary_df = []
        for line in info['summary_lines']:
            summary_df.append({
                '课程': line['session'],
                '上课': line['上课'],
                '旷课': line['旷课'],
                '病假': line['病假'],
                '事假': line['事假'],
            })
        st.dataframe(summary_df, use_container_width=True, hide_index=True)

    st.divider()
    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        st.download_button(
            label="📥 下载考勤明细 Excel",
            data=buf,
            file_name="学生考勤明细表.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    with col_dl2:
        st.download_button(
            label="📋 下载过程性成绩记载表",
            data=info['process_score_buf'],
            file_name="过程性成绩记载表.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# ── 单文件模式 ──
if mode == "单文件模式":
    uploaded = st.file_uploader("选择雨课堂导出的 Excel 文件（仅支持 .xlsx 格式）")

    workbook_source = None
    source_label = None
    if uploaded:
        if not uploaded.name.endswith('.xlsx'):
            st.error(f'不支持的文件格式："{uploaded.name}"，请上传 .xlsx 文件。')
        else:
            try:
                workbook_source = load_workbook(uploaded, data_only=True)
                source_label = "upload"
                st.session_state.show_demo = False
            except Exception:
                st.error("无法读取该文件，请确认上传的是有效的 .xlsx 格式文件。")
    elif st.session_state.get("show_demo"):
        demo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), DEMO_FILE)
        if os.path.exists(demo_path):
            try:
                workbook_source = load_workbook(demo_path, data_only=True)
                source_label = "demo"
            except Exception:
                st.error("无法读取示例文件。")
                st.session_state.show_demo = False

    if not uploaded and not st.session_state.get("show_demo"):
        if st.button("💡 加载示例数据看看效果", use_container_width=True):
            st.session_state.show_demo = True
            st.rerun()

    # ── Process single file ──
    if workbook_source:
        if source_label == "demo":
            st.info("当前展示的是示例数据，你可以上传自己的文件替换。")

        try:
            status = st.status("正在解析...", expanded=True)
            with status:
                st.write("读取 Excel 文件...")
                wb = workbook_source
                session_keys, session_sign_map, session_score_map, students, leave_data = parse_file(wb)
                wb.close()
                st.write("生成输出文件...")
                buf, info = generate_output(session_keys, session_sign_map, session_score_map, students, leave_data)
            status.update(label="解析完成", state="complete", expanded=False)
            show_results(buf, info)
        except ValueError as e:
            st.error(str(e))
        except Exception:
            st.error("解析过程出现未知错误，请确认上传的是雨课堂批量导出的汇总数据表（.xlsx）。")
    else:
        st.info("请上传文件开始分析。")

# ── 合并模式 ──
elif mode == "合并模式":
    st.caption("请确保两个文件中的学生姓名和学号一致，否则合并结果会不准确。")
    col1, col2 = st.columns(2)
    with col1:
        f1 = st.file_uploader("选择文件一", key="f1")
    with col2:
        f2 = st.file_uploader("选择文件二", key="f2")

    if f1 and f2:
        errors = []
        if not f1.name.endswith('.xlsx'):
            errors.append(f'文件一格式不支持："{f1.name}"')
        if not f2.name.endswith('.xlsx'):
            errors.append(f'文件二格式不支持："{f2.name}"')
        if errors:
            st.error('\n'.join(errors))
            st.stop()

        try:
            status = st.status("正在解析...", expanded=True)
            with status:
                st.write("解析文件一...")
                wb1 = load_workbook(f1, data_only=True)
                sk1, ssm1, sscm1, students1, ld1 = parse_file(wb1)
                wb1.close()

                st.write("解析文件二...")
                wb2 = load_workbook(f2, data_only=True)
                sk2, ssm2, sscm2, students2, ld2 = parse_file(wb2)
                wb2.close()

                st.write("验证学生信息一致性...")
                ids1 = [(s['id'], s['name']) for s in students1]
                ids2 = [(s['id'], s['name']) for s in students2]
                if ids1 != ids2:
                    set1, set2 = set(ids1), set(ids2)
                    diff = []
                    for s in set1 - set2:
                        diff.append(f"  文件一有但文件二缺少：学号 {s[0]} {s[1]}")
                    for s in set2 - set1:
                        diff.append(f"  文件二有但文件一缺少：学号 {s[0]} {s[1]}")
                    raise ValueError("两个文件的学生名单不一致：\n" + "\n".join(diff))

                st.write("合并数据...")
                session_keys = sk1 + sk2
                session_sign_map = OrderedDict(list(ssm1.items()) + list(ssm2.items()))
                session_score_map = OrderedDict(list(sscm1.items()) + list(sscm2.items()))
                students = []
                for s1, s2 in zip(students1, students2):
                    students.append({
                        'id': s1['id'],
                        'name': s1['name'],
                        'cls': s1['cls'],
                        'attendance': {**s1['attendance'], **s2['attendance']},
                        'scores': {**s1['scores'], **s2['scores']},
                    })
                leave_data = {**ld1, **ld2}

                st.write("生成输出文件...")
                buf, info = generate_output(session_keys, session_sign_map, session_score_map, students, leave_data)
            status.update(label="解析完成", state="complete", expanded=False)
            show_results(buf, info)
        except ValueError as e:
            st.error(str(e))
        except Exception:
            st.error("解析过程出现未知错误，请确认上传的是雨课堂批量导出的汇总数据表（.xlsx）。")
    else:
        st.info("请上传两个考勤文件开始合并分析。")
