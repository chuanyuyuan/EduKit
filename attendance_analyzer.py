#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
雨课堂考勤分析 — CLI 版
用法:
  python attendance_analyzer.py <文件.xlsx>
  python attendance_analyzer.py <文件一.xlsx> <文件二.xlsx>
"""

import sys
import os
import time
import argparse
from collections import OrderedDict
from io import BytesIO

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def col_idx(col):
    i = 0
    for ch in col:
        i = i * 26 + (ord(ch) - 64)
    return i - 1


def parse_summary(wb):
    if not wb.sheetnames:
        raise ValueError('Excel 文件中没有找到任何工作表。')

    ws = wb[wb.sheetnames[0]]

    row2 = {}
    for cell in ws[2]:
        if cell.value:
            row2[cell.column_letter] = str(cell.value)

    if not row2:
        raise ValueError('汇总表第 2 行为空，未找到子表头信息。请确认上传的是雨课堂批量导出文件（含汇总页）。')

    row1 = {}
    for cell in ws[1]:
        if cell.value:
            row1[cell.column_letter] = str(cell.value)

    if not row1:
        raise ValueError('汇总表第 1 行为空，未找到课堂名称合并表头。')

    sign_cols = sorted(
        [c for c, v in row2.items() if v == '签到方式'],
        key=col_idx
    )
    if not sign_cols:
        raise ValueError('未找到"签到方式"列。请确认上传的是雨课堂导出的"汇总-数据表"文件。')

    score_cols = sorted(
        [c for c, v in row2.items() if v.startswith('得分')],
        key=col_idx
    )
    session_headers = sorted(row1.keys(), key=col_idx)

    def map_cols_to_sessions(cols):
        result = OrderedDict()
        keys = []
        for c in cols:
            ci = col_idx(c)
            nearest = None
            nearest_i = -1
            for sh in session_headers:
                hi = col_idx(sh)
                if hi <= ci and hi > nearest_i:
                    nearest = row1[sh]
                    nearest_i = hi
            if nearest:
                result[nearest] = c
                keys.append(nearest)
        return result, keys

    session_sign_map, session_keys = map_cols_to_sessions(sign_cols)
    session_score_map, _ = map_cols_to_sessions(score_cols)

    students = []
    for row in ws.iter_rows(min_row=3, values_only=False):
        cells = {}
        for c in row:
            if c.value is None:
                continue
            try:
                cells[c.column_letter] = str(c.value)
            except AttributeError:
                pass
        sid = cells.get('A', '')
        name = cells.get('D', '')
        if not sid and not name:
            continue

        rec = {
            'id': sid,
            'name': name,
            'dept': cells.get('B', ''),
            'cls': cells.get('C', ''),
            'attendance': {},
            'scores': {},
        }
        for sk, sc in session_sign_map.items():
            rec['attendance'][sk] = cells.get(sc, '')
        for sk, sc in session_score_map.items():
            raw = cells.get(sc, '')
            try:
                rec['scores'][sk] = int(float(raw)) if raw else None
            except ValueError:
                rec['scores'][sk] = None
        students.append(rec)

    return session_keys, session_sign_map, session_score_map, students


def parse_sub_sheets(wb, session_keys):
    leave_data = {}
    idx = 0
    found = False

    for name in wb.sheetnames:
        if '课堂情况' not in name:
            continue
        found = True
        if idx >= len(session_keys):
            break

        session_key = session_keys[idx]
        idx += 1
        ws = wb[name]

        for row in ws.iter_rows(min_row=4, values_only=False):
            cells = {}
            for c in row:
                if c.value is None:
                    continue
                try:
                    cells[c.column_letter] = str(c.value)
                except AttributeError:
                    pass
            sid = cells.get('A', '')
            remark = cells.get('G', '')
            if sid and remark in ('病假', '事假'):
                leave_data[(session_key, sid)] = remark

    if not found:
        raise ValueError('未找到名称包含"课堂情况"的子表。请确认上传的是雨课堂批量导出文件（含每次课的课堂情况子表）。')

    return leave_data


def parse_single_session(wb):
    """Parse a single-session export file (no summary page, just one 课堂情况 sheet)."""
    ws = wb[wb.sheetnames[0]]

    raw_name = str(ws.cell(1, 1).value or '').strip()
    if not raw_name:
        raise ValueError('文件第 1 行为空，未找到课堂名称。')

    row3_col5 = str(ws.cell(3, 5).value or '')
    if '签到方式' not in row3_col5:
        raise ValueError('第 3 行未找到"签到方式"列，请确认是雨课堂导出的课堂数据表。')

    students = []
    for row_idx in range(5, ws.max_row + 1):
        sid = str(ws.cell(row_idx, 1).value or '').strip()
        name = str(ws.cell(row_idx, 4).value or '').strip()
        if not sid and not name:
            continue
        students.append({
            'id': sid,
            'name': name,
            'dept': '',
            'cls': '',
            'attendance': {},
            'scores': {},
        })
        students[-1]['_sign'] = str(ws.cell(row_idx, 5).value or '')
        students[-1]['_remark'] = str(ws.cell(row_idx, 7).value or '')

    session_keys = [raw_name]
    session_sign_map = OrderedDict([(raw_name, raw_name)])
    session_score_map = OrderedDict()

    leave_data = {}
    for s in students:
        s['attendance'][raw_name] = s.pop('_sign')
        remark = s.pop('_remark')
        if remark in ('病假', '事假'):
            leave_data[(raw_name, s['id'])] = remark

    return session_keys, session_sign_map, session_score_map, students, leave_data


def parse_file(wb):
    """Auto-detect file type (full export or single session) and parse."""
    if len(wb.sheetnames) == 1 and '课堂情况' in wb.sheetnames[0]:
        return parse_single_session(wb)
    sk, ssm, sscm, students = parse_summary(wb)
    ld = parse_sub_sheets(wb, sk)
    return sk, ssm, sscm, students, ld


PRESENT_SET = {'扫二维码', '“正在上课”提示', '教师添加', '课堂暗号'}


def generate_output(session_keys, session_sign_map, session_score_map, students, leave_data):
    """Generate output Excel bytes + summary info from parsed data."""

    session_count = len(session_sign_map)

    # ── Build data for Sheet 1: 考勤明细 ──
    out_rows = []
    for s in students:
        row = [s['id'], s['name']]
        absent_cnt = 0
        excused_cnt = 0
        for sk in session_sign_map:
            raw = s['attendance'].get(sk, '')
            if raw in PRESENT_SET:
                status = '上课'
            elif raw == '未上课':
                leave = leave_data.get((sk, s['id']))
                if leave == '病假':
                    status = '病假'
                    excused_cnt += 1
                elif leave == '事假':
                    status = '事假'
                    excused_cnt += 1
                else:
                    status = '旷课'
                    absent_cnt += 1
            else:
                status = raw
            row.append(status)
        row.append(f'{absent_cnt / session_count:.0%}')
        row.append(f'{(absent_cnt + excused_cnt) / session_count:.0%}')
        out_rows.append(row)

    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    color_map = {'上课': green_fill, '病假': yellow_fill, '事假': yellow_fill, '旷课': red_fill}
    gold_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')

    owb = Workbook()
    ows = owb.active
    ows.title = '考勤明细'
    headers = ['学号', '姓名'] + list(session_sign_map.keys()) + ['无故旷课率', '总旷课率']
    ows.append(headers)
    for row in out_rows:
        ows.append(row)
    for row in ows.iter_rows(min_row=2, max_row=ows.max_row,
                             min_col=3, max_col=2 + session_count):
        for cell in row:
            fill = color_map.get(cell.value)
            if fill:
                cell.fill = fill

    # ── Sheet 2: 课堂表现 ──
    ows2 = owb.create_sheet('课堂表现')
    score_headers = ['学号', '姓名'] + [k for k in session_score_map.keys()] + ['总分']
    ows2.append(score_headers)
    totals = []
    for s in students:
        row = [s['id'], s['name']]
        total = 0
        for sk in session_score_map:
            v = s['scores'].get(sk)
            if v is not None:
                row.append(v)
                total += v
            else:
                row.append('')
        row.append(total)
        totals.append(total)
        ows2.append(row)

    sorted_totals = sorted(totals, reverse=True)
    n_top = max(1, round(len(sorted_totals) * 0.1))
    threshold = sorted_totals[n_top - 1] if n_top <= len(sorted_totals) else sorted_totals[-1]
    for row in ows2.iter_rows(min_row=2, max_row=ows2.max_row):
        if row[-1].value is not None and row[-1].value >= threshold:
            for cell in row:
                cell.fill = gold_fill

    buf = BytesIO()
    owb.save(buf)
    buf.seek(0)

    # ── Summary stats ──
    summary_lines = []
    for sk in session_sign_map:
        attend = absent = sick = personal = 0
        for s in students:
            raw = s['attendance'].get(sk, '')
            if raw in PRESENT_SET:
                attend += 1
            elif raw == '未上课':
                leave = leave_data.get((sk, s['id']))
                if leave == '病假':
                    sick += 1
                elif leave == '事假':
                    personal += 1
                else:
                    absent += 1
        summary_lines.append({
            'session': sk,
            '上课': attend,
            '旷课': absent,
            '病假': sick,
            '事假': personal,
        })

    total_absent = sum(1 for s in students for sk in session_sign_map
                       if s['attendance'].get(sk) == '未上课'
                       and (sk, s['id']) not in leave_data)
    total_sick = sum(1 for v in leave_data.values() if v == '病假')
    total_personal = sum(1 for v in leave_data.values() if v == '事假')

    return buf, {
        'students': students,
        'session_count': session_count,
        'total_absent': total_absent,
        'total_sick': total_sick,
        'total_personal': total_personal,
        'summary_lines': summary_lines,
        'session_keys': session_keys,
        'leave_data': leave_data,
    }


def generate_process_score_sheet(students, session_keys, leave_data):
    """生成过程性成绩记载表（✓/✗/△），与样表格式一致"""
    owb = Workbook()
    ows = owb.active
    ows.title = '过程性成绩记载表'

    n_sessions = len(session_keys)
    session_start = 7
    score_col = max(session_start + n_sessions, 26)
    remark_col = score_col + 1

    headers = ['序号', '学号', '姓名', '性别', '专业', '班级'] + list(session_keys)
    while len(headers) < score_col - 1:
        headers.append('')
    headers.append('成绩')
    headers.append('备注')
    ows.append(headers)

    for i, s in enumerate(students, start=1):
        row = [i, s['id'], s['name'], '', '', s.get('cls', '')]
        for sk in session_keys:
            raw = s['attendance'].get(sk, '')
            if raw in PRESENT_SET:
                row.append('✓')
            elif raw == '未上课':
                leave = leave_data.get((sk, s['id']))
                row.append('△' if leave in ('病假', '事假') else '✗')
            else:
                row.append(raw)
        while len(row) < score_col - 1:
            row.append('')
        row.append('')
        row.append('')
        ows.append(row)

    font_song = Font(name='SimSun', size=9)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for row in ows.iter_rows(min_row=1, max_row=ows.max_row,
                              min_col=1, max_col=remark_col):
        for cell in row:
            cell.font = font_song
            cell.alignment = align_center
            cell.border = thin_border

    col_widths = {
        'A': 2.85, 'B': 12.75, 'C': 11.26, 'D': 3.25,
        'E': 15.87, 'F': 15.47,
    }
    for col_letter, width in col_widths.items():
        ows.column_dimensions[col_letter].width = width
    for i in range(session_start, score_col):
        ows.column_dimensions[get_column_letter(i)].width = 3.69
    ows.column_dimensions[get_column_letter(score_col)].width = 5.38
    ows.column_dimensions[get_column_letter(remark_col)].width = 9.77

    ows.row_dimensions[1].height = 22.6
    for r in range(2, ows.max_row + 1):
        ows.row_dimensions[r].height = 14.3

    buf = BytesIO()
    owb.save(buf)
    buf.seek(0)
    return buf


def main():
    parser = argparse.ArgumentParser(
        description='雨课堂考勤数据分析工具 — CLI 版',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
支持的文件类型:
  批量导出的汇总数据表（含汇总页 + 课堂情况子表）
  单次课导出的课堂数据表（自动识别）

输出文件:
  考勤明细_时间戳.xlsx      — 考勤明细（颜色标注）+ 课堂表现得分
  过程性成绩_时间戳.xlsx     — 过程性成绩记载表（✓/✗/△ 符号）
        """)
    parser.add_argument('input', nargs='+',
                        help='输入文件路径（1 个 = 单文件模式, 2 个 = 合并模式）')
    args = parser.parse_args()

    if len(args.input) == 0:
        parser.print_help()
        sys.exit(1)

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    dir_path = os.path.dirname(os.path.abspath(__file__))

    for f in args.input:
        if not os.path.exists(f):
            print(f"错误：文件不存在 — {f}")
            sys.exit(1)

    if len(args.input) == 1:
        # ── 单文件模式 ──
        input_path = args.input[0]
        print(f"解析文件: {input_path}")

        try:
            wb = load_workbook(input_path, data_only=True)
            session_keys, session_sign_map, session_score_map, students, leave_data = parse_file(wb)
            wb.close()
        except ValueError as e:
            print(f"解析错误：{e}")
            sys.exit(1)

        print(f"  共 {len(students)} 名学生, {len(session_sign_map)} 次课")

    elif len(args.input) >= 2:
        # ── 合并模式 ──
        path1, path2 = args.input[:2]
        print(f"解析文件一: {path1}")

        try:
            wb1 = load_workbook(path1, data_only=True)
            sk1, ssm1, sscm1, students1, ld1 = parse_file(wb1)
            wb1.close()
        except ValueError as e:
            print(f"解析文件一错误：{e}")
            sys.exit(1)

        print(f"  共 {len(students1)} 名学生, {len(sk1)} 次课")

        try:
            wb2 = load_workbook(path2, data_only=True)
            sk2, ssm2, sscm2, students2, ld2 = parse_file(wb2)
            wb2.close()
        except ValueError as e:
            print(f"解析文件二错误：{e}")
            sys.exit(1)

        print(f"\n解析文件二: {path2}")
        print(f"  共 {len(students2)} 名学生, {len(sk2)} 次课")

        # 校验一致性
        ids1 = [(s['id'], s['name']) for s in students1]
        ids2 = [(s['id'], s['name']) for s in students2]
        if ids1 != ids2:
            set1, set2 = set(ids1), set(ids2)
            diff = [f"  文件一有但文件二缺少：学号 {s[0]} {s[1]}" for s in set1 - set2]
            diff += [f"  文件二有但文件一缺少：学号 {s[0]} {s[1]}" for s in set2 - set1]
            print(f"错误：两个文件的学生名单不一致：\n" + "\n".join(diff))
            sys.exit(1)

        session_keys = sk1 + sk2
        session_sign_map = OrderedDict(list(ssm1.items()) + list(ssm2.items()))
        session_score_map = OrderedDict(list(sscm1.items()) + list(sscm2.items()))
        students = []
        for s1, s2 in zip(students1, students2):
            students.append({
                'id': s1['id'],
                'name': s1['name'],
                'cls': s1['cls'],
                'attendance': {**s1['attendance'], **s2['attendance']},
                'scores': {**s1['scores'], **s2['scores']},
            })
        leave_data = {**ld1, **ld2}
        print(f"\n合并后: {len(students)} 名学生, {len(session_keys)} 次课")

    # ── 生成输出 ──
    buf, info = generate_output(
        session_keys, session_sign_map, session_score_map, students, leave_data
    )

    out_path = os.path.join(dir_path, f'考勤明细_{timestamp}.xlsx')
    with open(out_path, 'wb') as f:
        f.write(buf.read())
    print(f"\n已生成: {os.path.basename(out_path)}")

    # ── 过程性成绩记载表 ──
    ps_buf = generate_process_score_sheet(students, session_keys, leave_data)
    ps_path = os.path.join(dir_path, f'过程性成绩记载表_{timestamp}.xlsx')
    with open(ps_path, 'wb') as f:
        f.write(ps_buf.read())
    print(f"已生成: {os.path.basename(ps_path)}")

    # ── 输出统计 ──
    print(f"\n统计摘要:")
    print(f"  学生人数: {len(students)}")
    print(f"  课堂次数: {info['session_count']}")
    for line in info['summary_lines']:
        print(f"  {line['session']:20s}  上课{line['上课']:>2}人  "
              f"旷课{line['旷课']:>2}人  病假{line['病假']:>2}人  "
              f"事假{line['事假']:>2}人")
    print(f"\n  旷课总人次: {info['total_absent']}")
    print(f"  病假总人次: {info['total_sick']}")
    print(f"  事假总人次: {info['total_personal']}")


if __name__ == '__main__':
    main()
