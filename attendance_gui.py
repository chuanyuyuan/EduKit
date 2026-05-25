#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
雨课堂考勤分析 - 本地图形界面版
tkinter 实现，无需联网，可打包为 exe
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import sys
import time
from datetime import datetime
from collections import OrderedDict
from io import BytesIO

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 常量 ──
PRESENT_SET = {'扫二维码', '“正在上课”提示', '教师添加', '课堂暗号'}


# ════════════════════════════════════════════════════════════
# 核心解析函数
# ════════════════════════════════════════════════════════════

def col_idx(col):
    i = 0
    for ch in col:
        i = i * 26 + (ord(ch) - 64)
    return i - 1


def parse_summary(wb):
    if not wb.sheetnames:
        raise ValueError('Excel 文件中没有找到任何工作表。')

    ws = wb[wb.sheetnames[0]]

    row2 = {}
    for cell in ws[2]:
        if cell.value:
            row2[cell.column_letter] = str(cell.value)

    if not row2:
        raise ValueError('汇总表第 2 行为空，未找到子表头信息。请确认上传的是雨课堂批量导出文件（含汇总页）。')

    row1 = {}
    for cell in ws[1]:
        if cell.value:
            row1[cell.column_letter] = str(cell.value)

    if not row1:
        raise ValueError('汇总表第 1 行为空，未找到课堂名称合并表头。')

    sign_cols = sorted(
        [c for c, v in row2.items() if v == '签到方式'],
        key=col_idx
    )
    if not sign_cols:
        raise ValueError('未找到"签到方式"列。请确认上传的是雨课堂导出的"汇总-数据表"文件。')

    score_cols = sorted(
        [c for c, v in row2.items() if v.startswith('得分')],
        key=col_idx
    )
    session_headers = sorted(row1.keys(), key=col_idx)

    def map_cols_to_sessions(cols):
        result = OrderedDict()
        keys = []
        for c in cols:
            ci = col_idx(c)
            nearest = None
            nearest_i = -1
            for sh in session_headers:
                hi = col_idx(sh)
                if hi <= ci and hi > nearest_i:
                    nearest = row1[sh]
                    nearest_i = hi
            if nearest:
                result[nearest] = c
                keys.append(nearest)
        return result, keys

    session_sign_map, session_keys = map_cols_to_sessions(sign_cols)
    session_score_map, _ = map_cols_to_sessions(score_cols)

    students = []
    for row in ws.iter_rows(min_row=3, values_only=False):
        cells = {}
        for c in row:
            if c.value is None:
                continue
            try:
                cells[c.column_letter] = str(c.value)
            except AttributeError:
                pass
        sid = cells.get('A', '')
        name = cells.get('D', '')
        if not sid and not name:
            continue

        rec = {
            'id': sid,
            'name': name,
            'dept': cells.get('B', ''),
            'cls': cells.get('C', ''),
            'attendance': {},
            'scores': {},
        }
        for sk, sc in session_sign_map.items():
            rec['attendance'][sk] = cells.get(sc, '')
        for sk, sc in session_score_map.items():
            raw = cells.get(sc, '')
            try:
                rec['scores'][sk] = int(float(raw)) if raw else None
            except ValueError:
                rec['scores'][sk] = None
        students.append(rec)

    return session_keys, session_sign_map, session_score_map, students


def parse_sub_sheets(wb, session_keys):
    leave_data = {}
    idx = 0
    found = False

    for name in wb.sheetnames:
        if '课堂情况' not in name:
            continue
        found = True
        if idx >= len(session_keys):
            break

        session_key = session_keys[idx]
        idx += 1
        ws = wb[name]

        for row in ws.iter_rows(min_row=4, values_only=False):
            cells = {}
            for c in row:
                if c.value is None:
                    continue
                try:
                    cells[c.column_letter] = str(c.value)
                except AttributeError:
                    pass
            sid = cells.get('A', '')
            remark = cells.get('G', '')
            if sid and remark in ('病假', '事假'):
                leave_data[(session_key, sid)] = remark

    if not found:
        raise ValueError('未找到名称包含"课堂情况"的子表。请确认上传的是雨课堂批量导出文件。')

    return leave_data


def parse_single_session(wb):
    ws = wb[wb.sheetnames[0]]

    raw_name = str(ws.cell(1, 1).value or '').strip()
    if not raw_name:
        raise ValueError('文件第 1 行为空，未找到课堂名称。')

    row3_col5 = str(ws.cell(3, 5).value or '')
    if '签到方式' not in row3_col5:
        raise ValueError('第 3 行未找到"签到方式"列，请确认是雨课堂导出的课堂数据表。')

    students = []
    for row_idx in range(5, ws.max_row + 1):
        sid = str(ws.cell(row_idx, 1).value or '').strip()
        name = str(ws.cell(row_idx, 4).value or '').strip()
        if not sid and not name:
            continue
        students.append({
            'id': sid,
            'name': name,
            'dept': '',
            'cls': '',
            'attendance': {},
            'scores': {},
        })
        students[-1]['_sign'] = str(ws.cell(row_idx, 5).value or '')
        students[-1]['_remark'] = str(ws.cell(row_idx, 7).value or '')

    session_keys = [raw_name]
    session_sign_map = OrderedDict([(raw_name, raw_name)])
    session_score_map = OrderedDict()

    leave_data = {}
    for s in students:
        s['attendance'][raw_name] = s.pop('_sign')
        remark = s.pop('_remark')
        if remark in ('病假', '事假'):
            leave_data[(raw_name, s['id'])] = remark

    return session_keys, session_sign_map, session_score_map, students, leave_data


def parse_file(wb):
    if len(wb.sheetnames) == 1 and '课堂情况' in wb.sheetnames[0]:
        return parse_single_session(wb)
    sk, ssm, sscm, students = parse_summary(wb)
    ld = parse_sub_sheets(wb, sk)
    return sk, ssm, sscm, students, ld


# ════════════════════════════════════════════════════════════
# 输出文件生成
# ════════════════════════════════════════════════════════════

def generate_output(session_keys, session_sign_map, session_score_map, students, leave_data):
    """生成考勤明细 Excel，返回 (bytes, info_dict)"""
    session_count = len(session_sign_map)

    out_rows = []
    for s in students:
        row = [s['id'], s['name']]
        absent_cnt = 0
        excused_cnt = 0
        for sk in session_sign_map:
            raw = s['attendance'].get(sk, '')
            if raw in PRESENT_SET:
                status = '上课'
            elif raw == '未上课':
                leave = leave_data.get((sk, s['id']))
                if leave == '病假':
                    status = '病假'
                    excused_cnt += 1
                elif leave == '事假':
                    status = '事假'
                    excused_cnt += 1
                else:
                    status = '旷课'
                    absent_cnt += 1
            else:
                status = raw
            row.append(status)
        row.append(f'{absent_cnt / session_count:.0%}')
        row.append(f'{(absent_cnt + excused_cnt) / session_count:.0%}')
        out_rows.append(row)

    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    color_map = {'上课': green_fill, '病假': yellow_fill, '事假': yellow_fill, '旷课': red_fill}
    gold_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')

    owb = Workbook()
    ows = owb.active
    ows.title = '考勤明细'
    headers = ['学号', '姓名'] + list(session_sign_map.keys()) + ['无故旷课率', '总旷课率']
    ows.append(headers)
    for row in out_rows:
        ows.append(row)
    for row in ows.iter_rows(min_row=2, max_row=ows.max_row,
                              min_col=3, max_col=2 + session_count):
        for cell in row:
            fill = color_map.get(cell.value)
            if fill:
                cell.fill = fill

    # 课堂表现 sheet
    ows2 = owb.create_sheet('课堂表现')
    score_headers = ['学号', '姓名'] + [k for k in session_score_map.keys()] + ['总分']
    ows2.append(score_headers)
    totals = []
    for s in students:
        row = [s['id'], s['name']]
        total = 0
        for sk in session_score_map:
            v = s['scores'].get(sk)
            if v is not None:
                row.append(v)
                total += v
            else:
                row.append('')
        row.append(total)
        totals.append(total)
        ows2.append(row)

    sorted_totals = sorted(totals, reverse=True)
    n_top = max(1, round(len(sorted_totals) * 0.1))
    threshold = sorted_totals[n_top - 1] if n_top <= len(sorted_totals) else sorted_totals[-1]
    for row in ows2.iter_rows(min_row=2, max_row=ows2.max_row):
        if row[-1].value is not None and row[-1].value >= threshold:
            for cell in row:
                cell.fill = gold_fill

    buf = BytesIO()
    owb.save(buf)
    buf.seek(0)

    info = {
        'session_count': session_count,
        'total_absent': sum(1 for s in students for sk in session_sign_map
                            if s['attendance'].get(sk) == '未上课'
                            and (sk, s['id']) not in leave_data),
        'total_sick': sum(1 for v in leave_data.values() if v == '病假'),
        'total_personal': sum(1 for v in leave_data.values() if v == '事假'),
    }
    return buf, info


def generate_process_score_sheet(students, session_keys, leave_data):
    """生成过程性成绩记载表（✓/✗/△），与样表格式一致"""
    owb = Workbook()
    ows = owb.active
    ows.title = '过程性成绩记载表'

    n_sessions = len(session_keys)
    session_start = 7  # column G
    score_col = max(session_start + n_sessions, 26)
    remark_col = score_col + 1

    headers = ['序号', '学号', '姓名', '性别', '专业', '班级'] + list(session_keys)
    while len(headers) < score_col - 1:
        headers.append('')
    headers.append('成绩')
    headers.append('备注')
    ows.append(headers)

    for i, s in enumerate(students, start=1):
        row = [i, s['id'], s['name'], '', '', s['cls']]
        for sk in session_keys:
            raw = s['attendance'].get(sk, '')
            if raw in PRESENT_SET:
                row.append('✓')
            elif raw == '未上课':
                leave = leave_data.get((sk, s['id']))
                row.append('△' if leave in ('病假', '事假') else '✗')
            else:
                row.append(raw)
        while len(row) < score_col - 1:
            row.append('')
        row.append('')
        row.append('')
        ows.append(row)

    font_song = Font(name='SimSun', size=9)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for row in ows.iter_rows(min_row=1, max_row=ows.max_row, min_col=1, max_col=remark_col):
        for cell in row:
            cell.font = font_song
            cell.alignment = align_center
            cell.border = thin_border

    col_widths = {
        'A': 2.85, 'B': 12.75, 'C': 11.26, 'D': 3.25,
        'E': 15.87, 'F': 15.47,
    }
    for cl, w in col_widths.items():
        ows.column_dimensions[cl].width = w
    for i in range(session_start, score_col):
        ows.column_dimensions[get_column_letter(i)].width = 3.69
    ows.column_dimensions[get_column_letter(score_col)].width = 5.38
    ows.column_dimensions[get_column_letter(remark_col)].width = 9.77

    ows.row_dimensions[1].height = 22.6
    for r in range(2, ows.max_row + 1):
        ows.row_dimensions[r].height = 14.3

    buf = BytesIO()
    owb.save(buf)
    buf.seek(0)
    return buf


# ════════════════════════════════════════════════════════════
# tkinter GUI
# ════════════════════════════════════════════════════════════

class AttendanceGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("长江雨课堂考勤分析工具")
        self.root.geometry("720x500")
        self.root.minsize(620, 400)
        self.root.update_idletasks()
        # 居中显示
        w = self.root.winfo_width()
        h = self.root.winfo_height()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(f'+{(sw - w) // 2}+{(sh - h) // 2}')

        self.input_file1 = tk.StringVar()
        self.input_file2 = tk.StringVar()
        self.output_dir = tk.StringVar(value=os.getcwd())

        self._build_ui()
        self._build_menu()

    def _build_menu(self):
        menubar = tk.Menu(self.root)

        settings_menu = tk.Menu(menubar, tearoff=0)
        settings_menu.add_command(label="设置默认输出路径", command=self._set_default_dir)
        menubar.add_cascade(label="设置", menu=settings_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="下载示例表格", command=self._download_sample)
        help_menu.add_separator()
        help_menu.add_command(label="使用说明", command=self._show_help)
        help_menu.add_command(label="关于", command=self._show_about)
        menubar.add_cascade(label="帮助", menu=help_menu)

        self.root.config(menu=menubar)

    def _set_default_dir(self):
        path = filedialog.askdirectory(title="选择默认输出路径")
        if path:
            self.output_dir.set(path)

    def _download_sample(self):
        demo_file = '示例表格.xlsx'
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), demo_file)
        if not os.path.exists(path):
            messagebox.showerror("错误", f"未找到示例文件: {demo_file}")
            return
        save_path = filedialog.asksaveasfilename(
            title="保存示例表格",
            defaultextension=".xlsx",
            initialfile=demo_file,
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        if not save_path:
            return
        try:
            import shutil
            shutil.copy2(path, save_path)
            self._log_success(f'示例表格已保存到: {save_path}')
        except Exception as e:
            messagebox.showerror("错误", f"保存失败:\n{e}")

    def _show_help(self):
        msg = (
            "长江雨课堂考勤分析工具 - 使用说明\n\n"
            "1. 选择模式：单文件模式 / 合并模式\n"
            "2. 选择雨课堂导出的 Excel 文件\n"
            "3. 选择输出目录\n"
            "4. 点击「开始分析」\n\n"
            "支持的文件类型：\n"
            "  • 批量导出的汇总数据表（含汇总页 + 课堂情况子表）\n"
            "  • 单次课导出的课堂数据表（自动识别）\n\n"
            "输出文件：\n"
            "  • 考勤明细_时间戳.xlsx（颜色标注）\n"
            "  • 过程性成绩记载表_时间戳.xlsx（✓/✗/△）"
        )
        messagebox.showinfo("使用说明", msg)

    def _show_about(self):
        msg = (
            "长江雨课堂考勤分析工具\n"
            f"版本: v1.0.0\n\n"
            "基于雨课堂导出的 Excel 自动生成考勤统计。\n\n"
            "GitHub: https://github.com/chuanyuyuan/\n"
            "  RainClassroomAttendanceAnalyzer"
        )
        messagebox.showinfo("关于", msg)

    # ── UI 构建 ──

    def _build_ui(self):
        # Tab 切换
        self.notebook = ttk.Notebook(self.root)
        self.tab_single = ttk.Frame(self.notebook, padding=10)
        self.tab_merge = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(self.tab_single, text="单文件模式")
        self.notebook.add(self.tab_merge, text="合并模式")
        self.notebook.pack(fill=tk.X, padx=10, pady=(10, 0))

        self._build_single_tab()
        self._build_merge_tab()
        self._build_log_area()

    def _file_picker_row(self, parent, row, label, text_var, btn_text, btn_cmd):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky=tk.W, pady=3)
        entry = ttk.Entry(parent, textvariable=text_var, width=50)
        entry.grid(row=row, column=1, sticky=tk.EW, padx=(5, 5), pady=3)
        ttk.Button(parent, text=btn_text, command=btn_cmd).grid(row=row, column=2, padx=(0, 5), pady=3)
        parent.columnconfigure(1, weight=1)

    def _build_single_tab(self):
        self._file_picker_row(self.tab_single, 0, "输入文件:", self.input_file1,
                              "选择文件", lambda: self._pick_file(self.input_file1))
        self._file_picker_row(self.tab_single, 1, "输出目录:", self.output_dir,
                              "选择目录", self._pick_output_dir)
        ttk.Button(self.tab_single, text="开始分析", command=self._run_single
                   ).grid(row=2, column=1, pady=(8, 0))

    def _build_merge_tab(self):
        self._file_picker_row(self.tab_merge, 0, "文件一:", self.input_file1,
                              "选择文件", lambda: self._pick_file(self.input_file1))
        self._file_picker_row(self.tab_merge, 1, "文件二:", self.input_file2,
                              "选择文件", lambda: self._pick_file(self.input_file2))
        self._file_picker_row(self.tab_merge, 2, "输出目录:", self.output_dir,
                              "选择目录", self._pick_output_dir)
        ttk.Label(self.tab_merge,
                  text="请确保两个文件中的学生姓名和学号一致，否则合并结果会不准确。",
                  foreground="#555", font=("Microsoft YaHei", 9)
                  ).grid(row=3, column=0, columnspan=2, pady=(2, 0), sticky="w")
        ttk.Button(self.tab_merge, text="开始分析", command=self._run_merge
                   ).grid(row=4, column=1, pady=(8, 0))

    def _build_log_area(self):
        frame = ttk.LabelFrame(self.root, text="运行日志", padding=5)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(3, 8))

        self.log_text = tk.Text(frame, height=10, wrap=tk.WORD, state=tk.DISABLED,
                                 font=("Microsoft YaHei", 9))
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    # ── 文件选取 ──

    def _pick_file(self, var):
        path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if path:
            var.set(path)

    def _pick_output_dir(self):
        path = filedialog.askdirectory(title="选择输出目录")
        if path:
            self.output_dir.set(path)

    # ── 日志 ──

    def _log(self, msg, tag=None):
        def _do():
            self.log_text.configure(state=tk.NORMAL)
            if tag:
                self.log_text.insert(tk.END, msg + '\n', tag)
            else:
                self.log_text.insert(tk.END, msg + '\n')
            self.log_text.see(tk.END)
            self.log_text.configure(state=tk.DISABLED)
        self.root.after(0, _do)

    def _log_error(self, msg):
        self._log(f'✗ {msg}', 'error')

    def _log_success(self, msg):
        self._log(f'✓ {msg}', 'success')

    def _log_info(self, msg):
        self._log(f'  {msg}', 'info')

    def _clear_log(self):
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete('1.0', tk.END)
        self.log_text.configure(state=tk.DISABLED)

    # ── 处理逻辑 ──

    def _run_single(self):
        path = self.input_file1.get()
        if not path:
            messagebox.showwarning("提示", "请先选择输入文件。")
            return
        if not os.path.isfile(path):
            messagebox.showerror("错误", "输入文件不存在。")
            return
        out_dir = self._ensure_out_dir()
        if out_dir is None:
            return
        self._clear_log()
        self._log_info(f'输入文件: {path}')
        self._log_info(f'输出目录: {out_dir}')
        threading.Thread(target=self._process,
                         args=('single', path, None, out_dir),
                         daemon=True).start()

    def _run_merge(self):
        path1 = self.input_file1.get()
        path2 = self.input_file2.get()
        if not path1 or not path2:
            messagebox.showwarning("提示", "请选择两个输入文件。")
            return
        if not os.path.isfile(path1):
            messagebox.showerror("错误", f"文件一不存在: {path1}")
            return
        if not os.path.isfile(path2):
            messagebox.showerror("错误", f"文件二不存在: {path2}")
            return
        out_dir = self._ensure_out_dir()
        if out_dir is None:
            return
        self._clear_log()
        self._log_info(f'文件一: {path1}')
        self._log_info(f'文件二: {path2}')
        self._log_info(f'输出目录: {out_dir}')
        threading.Thread(target=self._process,
                         args=('merge', path1, path2, out_dir),
                         daemon=True).start()

    def _ensure_out_dir(self):
        d = self.output_dir.get().strip()
        if not d:
            messagebox.showwarning("提示", "请先选择输出目录。")
            return None
        if not os.path.isdir(d):
            try:
                os.makedirs(d, exist_ok=True)
            except Exception as e:
                messagebox.showerror("错误", f"无法创建输出目录:\n{e}")
                return None
        return d

    def _process(self, mode, path1, path2, out_dir):
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')

        try:
            # ── 解析文件 ──
            self._log_info('正在读取 Excel 文件...')

            if mode == 'single':
                wb = load_workbook(path1, data_only=True)
                session_keys, session_sign_map, session_score_map, students, leave_data = parse_file(wb)
                wb.close()
            else:
                wb1 = load_workbook(path1, data_only=True)
                sk1, ssm1, sscm1, students1, ld1 = parse_file(wb1)
                wb1.close()

                wb2 = load_workbook(path2, data_only=True)
                sk2, ssm2, sscm2, students2, ld2 = parse_file(wb2)
                wb2.close()

                # 校验学生名单一致性
                ids1 = [(s['id'], s['name']) for s in students1]
                ids2 = [(s['id'], s['name']) for s in students2]
                if ids1 != ids2:
                    set1, set2 = set(ids1), set(ids2)
                    diff = []
                    for s in set1 - set2:
                        diff.append(f'文件一有但文件二缺少：学号 {s[0]} {s[1]}')
                    for s in set2 - set1:
                        diff.append(f'文件二有但文件一缺少：学号 {s[0]} {s[1]}')
                    raise ValueError('两个文件的学生名单不一致：\n' + '\n'.join(diff))

                session_keys = sk1 + sk2
                session_sign_map = OrderedDict(list(ssm1.items()) + list(ssm2.items()))
                session_score_map = OrderedDict(list(sscm1.items()) + list(sscm2.items()))
                students = []
                for s1, s2 in zip(students1, students2):
                    students.append({
                        'id': s1['id'],
                        'name': s1['name'],
                        'dept': s1['dept'],
                        'cls': s1['cls'],
                        'attendance': {**s1['attendance'], **s2['attendance']},
                        'scores': {**s1['scores'], **s2['scores']},
                    })
                leave_data = {**ld1, **ld2}

            self._log_success(f'解析完成，共 {len(session_sign_map)} 次课，{len(students)} 名学生')

            # ── 生成考勤明细 ──
            self._log_info('正在生成考勤明细表...')
            buf, info = generate_output(session_keys, session_sign_map, session_score_map,
                                        students, leave_data)
            name_att = f'考勤明细_{ts}.xlsx'
            out_path_att = os.path.join(out_dir, name_att)
            with open(out_path_att, 'wb') as f:
                f.write(buf.getvalue())
            self._log_success(f'已生成: {name_att}')

            # ── 生成过程性成绩记载表 ──
            self._log_info('正在生成过程性成绩记载表...')
            buf2 = generate_process_score_sheet(students, session_keys, leave_data)
            name_score = f'过程性成绩记载表_{ts}.xlsx'
            out_path_score = os.path.join(out_dir, name_score)
            with open(out_path_score, 'wb') as f:
                f.write(buf2.getvalue())
            self._log_success(f'已生成: {name_score}')

            # 统计摘要
            self._log_info(f'旷课总人次: {info["total_absent"]}  '
                           f'病假总人次: {info["total_sick"]}  '
                           f'事假总人次: {info["total_personal"]}')
            self._log_info('全部完成！')

        except ValueError as e:
            self._log_error(str(e))
        except Exception as e:
            self._log_error(f'处理出错: {e}')


# ════════════════════════════════════════════════════════════
# 入口
# ════════════════════════════════════════════════════════════

def main():
    root = tk.Tk()
    app = AttendanceGUI(root)

    # 配置日志颜色标签
    app.log_text.tag_configure('error', foreground='red')
    app.log_text.tag_configure('success', foreground='green')
    app.log_text.tag_configure('info', foreground='#555')

    # 设置字体渲染（Windows 下清晰显示 ✓/✗/△）
    style = ttk.Style()
    style.theme_use('vista' if 'vista' in style.theme_names() else 'clam')

    root.mainloop()


if __name__ == '__main__':
    main()
