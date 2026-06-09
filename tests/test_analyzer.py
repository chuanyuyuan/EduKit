"""
考勤分析工具 (CLI 版) — 功能测试
=================================
测试场景（与 test_app.py / test_gui.py 一致）:
  1. 单表模式 — 单表
  2. 单表模式 — 汇总表
  3. 合并模式 — 单表 + 单表
  4. 合并模式 — 单表 + 汇总表
  5. 合并模式 — 汇总表 + 汇总表
  6. 学生名单不一致检测
  7. 输出格式验证
  8. parse_file 自动识别
  9. 输出内容验证
  10. 过程性成绩符号验证

运行： python tests/test_py
"""
import sys, os, glob
sys.stdout.reconfigure(encoding='utf-8')

# ── 从 tools.attendance.core 导入 ──
PROJECT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEST_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, PROJECT)

from tools.attendance.core import (
    col_idx, parse_file, generate_output, generate_process_score_sheet,
    parse_summary, parse_sub_sheets, parse_single_session,
)

from collections import OrderedDict
from io import BytesIO
from openpyxl import load_workbook
PASS = FAIL = 0

def check(cond, msg):
    global PASS, FAIL
    if cond: PASS += 1; print(f'  ✓ {msg}')
    else:    FAIL += 1; print(f'  ✗ {msg}')

def check_eq(a, b, msg):
    global PASS, FAIL
    if a == b: PASS += 1; print(f'  ✓ {msg} ({a})')
    else:      FAIL += 1; print(f'  ✗ {msg}: got {a}, expected {b}')

def section(name):
    print(f'\n{"="*60}\n  {name}\n{"="*60}')


# ── Load test files, classify by type ──
_single_files = []
_summary_files = []

for fpath in sorted(glob.glob(os.path.join(TEST_DIR, '*.xlsx'))):
    wb = load_workbook(fpath, data_only=True)
    if len(wb.sheetnames) == 1 and '课堂情况' in wb.sheetnames[0]:
        _single_files.append(fpath)
    elif len(wb.sheetnames) > 1:
        _summary_files.append(fpath)
    wb.close()

assert len(_single_files) >= 2, f'需要 ≥2 个单表文件, 找到 {len(_single_files)}'
assert len(_summary_files) >= 2, f'需要 ≥2 个汇总表文件, 找到 {len(_summary_files)}'

S1 = _single_files[0]
S2 = _single_files[1]
M1 = _summary_files[0]
M2 = _summary_files[1]

print(f'测试文件 (CLI 版):')
print(f'  单表1: {os.path.basename(S1)}')
print(f'  单表2: {os.path.basename(S2)}')
print(f'  汇总表1: {os.path.basename(M1)}')
print(f'  汇总表2: {os.path.basename(M2)}')


def parse(path):
    wb = load_workbook(path, data_only=True)
    try:
        return parse_file(wb)
    finally:
        wb.close()


def merge(stus1, stus2):
    """Merge two student lists by ID (order-independent)."""
    stus2_by_id = {s['id']: s for s in stus2}
    return [{
        'id': s1['id'], 'name': s1['name'], 'cls': s1['cls'],
        'attendance': {**s1['attendance'], **stus2_by_id[s1['id']]['attendance']},
        'scores': {**s1['scores'], **stus2_by_id[s1['id']]['scores']},
    } for s1 in stus1]


def run_merge(path1, path2):
    sk1, ssm1, sscm1, stus1, ld1 = parse(path1)
    sk2, ssm2, sscm2, stus2, ld2 = parse(path2)
    ids1 = [(s['id'], s['name']) for s in stus1]
    ids2 = [(s['id'], s['name']) for s in stus2]
    assert ids1 == ids2, 'Student lists must match'
    sk = sk1 + sk2
    ssm = OrderedDict(list(ssm1.items()) + list(ssm2.items()))
    sscm = OrderedDict(list(sscm1.items()) + list(sscm2.items()))
    stus = merge(stus1, stus2)
    ld = {**ld1, **ld2}
    return sk, ssm, sscm, stus, ld


# ════════════════════════════════════════════
# TEST 1: 单表模式 — 单表
# ════════════════════════════════════════════
section('Test 1: 单表模式 — 单表')

sk, ssm, sscm, stus, ld = parse(S1)
check_eq(len(sk), 1, '解析到 1 次课')
check_eq(len(stus), 10, '解析到 10 名学生')
for s in stus:
    check(len(s['attendance']) == 1, f'{s["name"]} 有 1 次课记录')

# ════════════════════════════════════════════
# TEST 2: 单表模式 — 汇总表
# ════════════════════════════════════════════
section('Test 2: 单表模式 — 汇总表')

sk, ssm, sscm, stus, ld = parse(M1)
check_eq(len(sk), 3, '解析到 3 次课')
check_eq(len(stus), 10, '解析到 10 名学生')
for s in stus:
    check(len(s['attendance']) == 3, f'{s["name"]} 有 3 次课记录')

# ════════════════════════════════════════════
# TEST 3: 合并 — 单表 + 单表
# ════════════════════════════════════════════
section('Test 3: 合并模式 — 单表 + 单表')

sk, ssm, sscm, stus, ld = run_merge(S1, S2)
check_eq(len(sk), 2, '合并后 2 次课')
check_eq(len(stus), 10, '合并后 10 名学生')
for s in stus:
    check(len(s['attendance']) == 2, f'{s["name"]} 有 2 次课记录')
buf, info = generate_output(sk, ssm, sscm, stus, ld)
check(isinstance(buf, BytesIO), '输出生成成功')

# ════════════════════════════════════════════
# TEST 4: 合并 — 单表 + 汇总表
# ════════════════════════════════════════════
section('Test 4: 合并模式 — 单表 + 汇总表')

sk, ssm, sscm, stus, ld = run_merge(S1, M1)
check_eq(len(sk), 4, '合并后 4 次课')
check_eq(len(stus), 10, '合并后 10 名学生')
buf, info = generate_output(sk, ssm, sscm, stus, ld)
check(isinstance(buf, BytesIO), '输出生成成功')

# ════════════════════════════════════════════
# TEST 5: 合并 — 汇总表 + 汇总表
# ════════════════════════════════════════════
section('Test 5: 合并模式 — 汇总表 + 汇总表')

sk, ssm, sscm, stus, ld = run_merge(M1, M2)
check_eq(len(sk), 6, '合并后 6 次课')
check_eq(len(stus), 10, '合并后 10 名学生')
buf, info = generate_output(sk, ssm, sscm, stus, ld)
check(isinstance(buf, BytesIO), '输出生成成功')

# ════════════════════════════════════════════
# TEST 6: 一致性校验
# ════════════════════════════════════════════
section('Test 6: 学生名单不一致检测')

_, _, _, stus1, _ = parse(S1)
_, _, _, stus2, _ = parse(M1)
bad = [dict(s) for s in stus1]
bad[0]['id'] = '99999999'
s1 = {(s['id'], s['name']) for s in bad}
s2 = {(s['id'], s['name']) for s in stus2}
diff = [f'文件一有但文件二缺少：学号 {s[0]} {s[1]}' for s in s1 - s2]
diff += [f'文件二有但文件一缺少：学号 {s[0]} {s[1]}' for s in s2 - s1]
check(len(diff) > 0, f'名单不匹配时检测到 {len(diff)} 处差异')
check(any('99999999' in d for d in diff), '差异中包含修改的学号')

# ════════════════════════════════════════════
# TEST 7: 输出格式验证
# ════════════════════════════════════════════
section('Test 7: 输出格式验证')

sk, ssm, sscm, stus, ld = parse(M1)
buf, info = generate_output(sk, ssm, sscm, stus, ld)

out_wb = load_workbook(buf, data_only=False)
ws = out_wb.active
headers = [str(ws.cell(1, c).value or '') for c in range(1, ws.max_column + 1)]
check('无故旷课率' in ''.join(headers[-3:]), '考勤明细有无故旷课率列')
check('总旷课率' in ''.join(headers[-3:]), '考勤明细有总旷课率列')
check('课堂表现' in out_wb.sheetnames[1], f'Sheet2 为课堂表现')
out_wb.close()

# 过程性成绩记载表
ps_buf = generate_process_score_sheet(stus, sk, ld)
check(isinstance(ps_buf, BytesIO), '过程性成绩表生成成功')
ps_wb = load_workbook(ps_buf, data_only=False)
ps_ws = ps_wb.active
ps_headers = [str(ps_ws.cell(1, c).value or '') for c in range(1, ps_ws.max_column + 1)]
for col in ['序号', '学号', '姓名', '班级']:
    check(col in ps_headers, f'过程性成绩包含列: {col}')
cell = ps_ws.cell(2, 1)
check(cell.font.name == 'SimSun', f'字体为宋体, 实际: {cell.font.name}')
check(cell.font.size == 9, f'字号 9pt, 实际: {cell.font.size}')
check(cell.alignment.horizontal == 'center', '居中对齐')
check(cell.border.left.style == 'thin', '细边框')
ps_wb.close()

# ════════════════════════════════════════════
# TEST 8: parse_file 自动识别
# ════════════════════════════════════════════
section('Test 8: parse_file 自动识别')

sk_s, ssm_s, sscm_s, stus_s, ld_s = parse(S1)
check_eq(len(sk_s), 1, '单表文件识别为 1 次课')

sk_m, ssm_m, sscm_m, stus_m, ld_m = parse(M1)
check_eq(len(sk_m), 3, '汇总表文件识别为 3 次课')
check(stus_s[0]['id'] == stus_m[0]['id'], '两文件学号一致')

# col_idx 工具函数
check_eq(col_idx('A'), 0, 'col_idx(A) = 0')
check_eq(col_idx('Z'), 25, 'col_idx(Z) = 25')
check_eq(col_idx('AA'), 26, 'col_idx(AA) = 26')

# ════════════════════════════════════════════
# TEST 9: 输出内容验证
# ════════════════════════════════════════════
section('Test 9: 输出内容验证')

sk, ssm, sscm, stus, ld = parse(M1)
buf, info = generate_output(sk, ssm, sscm, stus, ld)
out_wb = load_workbook(buf, data_only=False)
ws = out_wb.active

# 数据行数 = 10 学生 + 1 表头
check_eq(ws.max_row, 11, '考勤明细行数 = 11（含表头）')
check_eq(ws.max_column, 7, '考勤明细列数 = 7（学号+姓名+3次课+2比率）')

# 统计每种考勤状态的出现次数
from collections import Counter
status_cnt = Counter()
for r in range(2, ws.max_row + 1):
    for c in range(3, 6):  # 3 attendance columns
        v = str(ws.cell(r, c).value or '')
        status_cnt[v] += 1
check(status_cnt.get('上课', 0) >= 20, f'上课状态 ≥20 次 ({status_cnt.get("上课", 0)})')
check(status_cnt.get('旷课', 0) >= 3, f'旷课状态 ≥3 次 ({status_cnt.get("旷课", 0)})')
check_eq(sum(status_cnt.values()), 30, '考勤状态总数 = 30')

# 验证颜色填充
green = '00C6EFCE'
red = '00FFC7CE'
fill_ok = 0
for r in range(2, ws.max_row + 1):
    for c in range(3, 6):
        cell = ws.cell(r, c)
        if cell.value == '上课' and cell.fill.start_color.rgb == green:
            fill_ok += 1
        elif cell.value == '旷课' and cell.fill.start_color.rgb == red:
            fill_ok += 1
check(fill_ok >= 20, f'颜色标注正确 ≥20 个单元格 ({fill_ok})')

# 验证比率列
rate_col = 6
for r in range(2, ws.max_row + 1):
    v = str(ws.cell(r, rate_col).value or '')
    check(v.endswith('%'), f'第 {r} 行旷课率格式: {v}')
out_wb.close()

# ════════════════════════════════════════════
# TEST 10: 过程性成绩符号验证
# ════════════════════════════════════════════
section('Test 10: 过程性成绩符号验证')

ps_buf = generate_process_score_sheet(stus, sk, ld)
ps_wb = load_workbook(ps_buf, data_only=False)
ps_ws = ps_wb.active

check_eq(ps_ws.max_row, 11, '过程性成绩表行数 = 11')
check(ps_ws.max_column >= 10, f'过程性成绩表列数 ≥ 10 ({ps_ws.max_column})')

valid_symbols = {'✓', '✗', '△'}
symbol_cnt = Counter()
for r in range(2, ps_ws.max_row + 1):
    for c in range(7, 10):  # session columns (G, H, I)
        v = str(ps_ws.cell(r, c).value or '')
        if v in valid_symbols:
            symbol_cnt[v] += 1
check_eq(sum(symbol_cnt.values()), 30, f'符号总数 = 30 ({dict(symbol_cnt)})')
check(symbol_cnt.get('✓', 0) >= 20, f'✓ 出现 ≥20 次 ({symbol_cnt.get("✓", 0)})')
check(symbol_cnt.get('✗', 0) >= 3, f'✗ 出现 ≥3 次 ({symbol_cnt.get("✗", 0)})')

for r in range(2, ps_ws.max_row + 1):
    sid = str(ps_ws.cell(r, 2).value or '')
    name = str(ps_ws.cell(r, 3).value or '')
    check(sid.startswith('0423'), f'{name} 学号以 0423 开头: {sid}')

ps_wb.close()

# ════════════════════════════════════════════
# TEST 11: 合并模式 — 顺序无关
# ════════════════════════════════════════════
section('Test 11: 合并模式 — 顺序无关')

# 解析两个汇总表，把第二个反转顺序
wb1 = load_workbook(M1, data_only=True)
sk1, ssm1, sscm1, stus1, ld1 = parse_file(wb1)
wb1.close()

wb2 = load_workbook(M2, data_only=True)
sk2, ssm2, sscm2, stus2_orig, ld2 = parse_file(wb2)
wb2.close()

# 反转 students2 的顺序
stus2 = list(reversed(stus2_orig))

# 用新顺序走合并流程
ids1 = {(s['id'], s['name']) for s in stus1}
ids2_set = {(s['id'], s['name']) for s in stus2}
check(ids1 == ids2_set, '顺序不同但集合一致')

stus2_by_id = {s['id']: s for s in stus2}
merged = []
for s1 in stus1:
    s2 = stus2_by_id[s1['id']]
    merged.append({
        'id': s1['id'],
        'name': s1['name'],
        'cls': s1['cls'],
        'attendance': {**s1['attendance'], **s2['attendance']},
        'scores': {**s1['scores'], **s2['scores']},
    })
check_eq(len(merged), 10, '合并后 10 名学生')
check_eq(len(sk1) + len(sk2), 6, '合并后 6 次课')
check(merged[0]['id'] == stus1[0]['id'], '合并结果按文件一顺序排列')

# ════════════════════════════════════════════
# TEST 12: 姓名不一致检测
# ════════════════════════════════════════════
section('Test 12: 姓名不一致检测')

_, _, _, stus_a, _ = parse(S1)
_, _, _, stus_b, _ = parse(M1)
# 改文件一的第一个学生姓名
bad_name = [dict(s) for s in stus_a]
bad_name[0]['name'] = '张改'
name1_by_id = {s['id']: s['name'] for s in bad_name}
name_mismatch_ids = set()
diff = []
for s in stus_b:
    if s['id'] in name1_by_id and name1_by_id[s['id']] != s['name']:
        diff.append(f'学号 {s["id"]} 姓名不一致')
        name_mismatch_ids.add(s['id'])
check(len(diff) > 0, '姓名不一致时检测到差异')
check(not name_mismatch_ids - {'04230001'}, f'只有学号 04230001 被标记为姓名不一致 ({name_mismatch_ids})')

# ════════════════════════════════════════════
# TEST 13: parse_summary 异常路径
# ════════════════════════════════════════════
section('Test 13: parse_summary 异常路径')

from openpyxl import Workbook

# 13a: 第 2 行为空
wb = Workbook()
ws = wb.active
ws.cell(1, 1, value="Session 1")
try:
    parse_summary(wb)
    check(False, '13a 空第 2 行应抛出 ValueError')
except ValueError as e:
    check('第 2 行为空' in str(e), f'13a 空第 2 行提示正确')
wb.close()

# 13b: 第 1 行为空
wb = Workbook()
ws = wb.active
ws.cell(2, 1, value="签到方式")
try:
    parse_summary(wb)
    check(False, '13b 空第 1 行应抛出 ValueError')
except ValueError as e:
    check('第 1 行为空' in str(e), f'13b 空第 1 行提示正确')
wb.close()

# 13c: 没有签到方式列
wb = Workbook()
ws = wb.active
ws.cell(1, 1, value="Session 1")
ws.cell(2, 1, value="得分")
try:
    parse_summary(wb)
    check(False, '13c 无签到方式列应抛出 ValueError')
except ValueError as e:
    check('签到方式' in str(e), f'13c 无签到方式列提示正确')
wb.close()


# ════════════════════════════════════════════
# TEST 14: parse_sub_sheets 异常路径
# ════════════════════════════════════════════
section('Test 14: parse_sub_sheets 异常路径')

wb = Workbook()
ws = wb.active
ws.title = "汇总"  # 不含「课堂情况」
try:
    parse_sub_sheets(wb, ["session1"])
    check(False, '14 无课堂情况子表应抛出 ValueError')
except ValueError as e:
    check('课堂情况' in str(e), f'14 无课堂情况子表提示正确')
wb.close()


# ════════════════════════════════════════════
# TEST 15: parse_single_session 异常路径
# ════════════════════════════════════════════
section('Test 15: parse_single_session 异常路径')

# 15a: 第 1 行为空
wb = Workbook()
ws = wb.active
ws.cell(3, 5, value="签到方式")
try:
    parse_single_session(wb)
    check(False, '15a 空第 1 行应抛出 ValueError')
except ValueError as e:
    check('第 1 行为空' in str(e), f'15a 空第 1 行提示正确')
wb.close()

# 15b: 无签到方式列
wb = Workbook()
ws = wb.active
ws.cell(1, 1, value="课堂名称")
ws.cell(3, 5, value="得分")
try:
    parse_single_session(wb)
    check(False, '15b 无签到方式应抛出 ValueError')
except ValueError as e:
    check('签到方式' in str(e), f'15b 无签到方式提示正确')
wb.close()


# ════════════════════════════════════════════
# TEST 16: parse_file 异常路径
# ════════════════════════════════════════════
section('Test 16: parse_file 异常路径')

# 16a: 空 Workbook（无 sheet）
wb = Workbook()
wb.remove(wb.active)
try:
    parse_file(wb)
    check(False, '16a 空表应抛出 ValueError')
except ValueError as e:
    check('工作表' in str(e), f'16a 空表提示正确')
wb.close()

# 16b: 空白 sheet 名（含课堂情况但无数据）
wb = Workbook()
ws = wb.active
ws.title = "课堂情况_1"
# parse_file → parse_single_session → no row1
try:
    parse_file(wb)
    check(False, '16b 空白单表应抛出 ValueError')
except ValueError as e:
    check(True, f'16b 空白单表正确抛出: {e}')
wb.close()


# ════════════════════════════════════════════
# TEST 17: 签到状态 — 未知值 fallthrough
# ════════════════════════════════════════════
section('Test 17: 签到状态 — 未知值 fallthrough')

# 当签到值既不在 PRESENT_SET 也不是「未上课」时，应原样保留
sk, ssm, sscm, stus, ld = parse(M1)
# 手动修改第一个学生的第一次课签到来模拟未知值
session_name = list(ssm.keys())[0]
stus[0]['attendance'][session_name] = '其他签到方式'
buf, info = generate_output(sk, ssm, sscm, stus, ld)
preview = info['preview_attendance']
check(preview[0][session_name] == '其他签到方式',
      f'17 未知签到值原样保留: {preview[0][session_name]}')


# ════════════════════════════════════════════
# TEST 18: 空 session_score_map
# ════════════════════════════════════════════
section('Test 18: 空 session_score_map')

buf, info = generate_output(sk, ssm, OrderedDict(), stus, ld)
check(isinstance(buf, BytesIO), '18a 空 score_map 输出生成成功')
check(info['session_count'] == len(ssm), '18b session_count 不受 score_map 影响')
if info['preview_scores']:
    check(info['preview_scores'][0]['总分'] == 0,
          '18c 无得分时总分为 0')


# ════════════════════════════════════════════
# TEST 19: 全班旷课 / 全班全勤
# ════════════════════════════════════════════
section('Test 19: 全班旷课 / 全班全勤')

# 把所有人的所有签到设为「未上课」且无请假
stus_all_absent = []
for s in stus:
    sa = dict(s)
    sa['attendance'] = {k: '未上课' for k in ssm}
    stus_all_absent.append(sa)

buf, info = generate_output(sk, ssm, sscm, stus_all_absent, {})
out_wb = load_workbook(buf, data_only=False)
ows = out_wb.active
absent_rate = str(ows.cell(2, ows.max_column - 1).value or '')
total_rate = str(ows.cell(2, ows.max_column).value or '')
check(absent_rate == '100%', f'19a 全班旷课率: {absent_rate}')
check(total_rate == '100%', f'19b 全班总旷课率: {total_rate}')
out_wb.close()

# 全班全勤
stus_all_present = []
for s in stus:
    sa = dict(s)
    sa['attendance'] = {k: '扫二维码' for k in ssm}
    stus_all_present.append(sa)

buf, info = generate_output(sk, ssm, sscm, stus_all_present, {})
out_wb = load_workbook(buf, data_only=False)
ows = out_wb.active
absent_rate = str(ows.cell(2, ows.max_column - 1).value or '')
check(absent_rate == '0%', f'19c 全班全勤旷课率: {absent_rate}')
out_wb.close()


# ════════════════════════════════════════════
# TEST 20: 课堂表现 — 分数相同时前 10% 高亮
# ════════════════════════════════════════════
section('Test 20: 分数相同时前 10% 高亮')

stus_same = []
for s in stus:
    sa = dict(s)
    sa['scores'] = {k: 80 for k in sscm}
    stus_same.append(sa)
buf, info = generate_output(sk, ssm, sscm, stus_same, {})
out_wb = load_workbook(buf, data_only=False)
ows2 = out_wb[out_wb.sheetnames[1]]
gold_count = 0
for row in ows2.iter_rows(min_row=2, max_row=ows2.max_row):
    v = row[-1].value  # 总分
    if v is not None:
        try:
            fill_rgb = row[-1].fill.start_color.rgb if row[-1].fill else ''
            if 'FFD700' in str(fill_rgb):
                gold_count += 1
        except AttributeError:
            pass
check(gold_count >= 1, f'20 同分时至少 1 行高亮（10% × 10人 = 1 行）({gold_count})')
out_wb.close()


# ════════════════════════════════════════════
# TEST 21: parse_single_session 正常路径
# ════════════════════════════════════════════
section('Test 21: parse_single_session 正常路径')

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "课堂情况_2024年3月5日"
ws.cell(1, 1, value="2024春_理论课")
ws.cell(3, 5, value="签到方式")
ws.cell(5, 1, value="04230001")
ws.cell(5, 4, value="赵一一")
ws.cell(5, 5, value="扫二维码")
ws.cell(5, 7, value="")
ws.cell(6, 1, value="04230002")
ws.cell(6, 4, value="钱二二")
ws.cell(6, 5, value="未上课")
ws.cell(6, 7, value="病假")
sk, ssm, sscm, students, ld = parse_single_session(wb)
check_eq(len(sk), 1, '21 解析到 1 次课')
check_eq(len(students), 2, '21 解析到 2 名学生')
check(students[0]['attendance'][sk[0]] == '扫二维码', '21 签到正确')
check(('2024春_理论课', '04230002') in ld, '21 请假记录存在')
check(ld[('2024春_理论课', '04230002')] == '病假', '21 请假类型正确')
wb.close()

# ════════════════════════════════════════════
section('SUMMARY')
print(f'  CLI 版: {PASS} 通过, {FAIL} 失败')
print(f'{"="*60}')
if FAIL:
    sys.exit(1)
