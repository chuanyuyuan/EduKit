"""
考勤分析工具 — 功能测试
========================
测试场景:
  1. 单表模式 — 单表（解析单次课导出文件）
  2. 单表模式 — 汇总表（解析完整批量导出文件）
  3. 合并模式 — 单表 + 单表
  4. 合并模式 — 单表 + 汇总表
  5. 合并模式 — 汇总表 + 汇总表

运行： python tests/test_py
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')

PROJECT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEST_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, PROJECT)

from tools.attendance.core import (
    col_idx, parse_file, generate_output, generate_process_score_sheet
)

from collections import OrderedDict
from io import BytesIO
from openpyxl import load_workbook
PASS = FAIL = 0

def check(cond, msg):
    global PASS, FAIL
    if cond: PASS += 1; print(f'  ✓ {msg}')
    else:    FAIL += 1; print(f'  ✗ {msg}')

def check_eq(a, b, msg):
    global PASS, FAIL
    if a == b: PASS += 1; print(f'  ✓ {msg} ({a})')
    else:      FAIL += 1; print(f'  ✗ {msg}: got {a}, expected {b}')

def section(name):
    print(f'\n{"="*60}\n  {name}\n{"="*60}')

# ── Load test files, classify by type ──
import glob
_single_files = []   # 单表 (1 sheet, 课堂情况 in name)
_summary_files = []  # 汇总表 (multiple sheets)

for fpath in sorted(glob.glob(os.path.join(TEST_DIR, '*.xlsx'))):
    wb = load_workbook(fpath, data_only=True)
    if len(wb.sheetnames) == 1 and '课堂情况' in wb.sheetnames[0]:
        _single_files.append(fpath)
    elif len(wb.sheetnames) > 1:
        _summary_files.append(fpath)
    wb.close()

assert len(_single_files) >= 2, f'需要 ≥2 个单表文件, 找到 {len(_single_files)}'
assert len(_summary_files) >= 2, f'需要 ≥2 个汇总表文件, 找到 {len(_summary_files)}'

S1 = _single_files[0]   # 单表测试表1 (0318实验课)
S2 = _single_files[1]   # 单表测试表2 (0305理论课)
M1 = _summary_files[0]  # 汇总表测试表1 (3月)
M2 = _summary_files[1]  # 汇总表测试表2 (4月)

print(f'测试文件:')
print(f'  单表1: {os.path.basename(S1)}')
print(f'  单表2: {os.path.basename(S2)}')
print(f'  汇总表1: {os.path.basename(M1)}')
print(f'  汇总表2: {os.path.basename(M2)}')


def parse(path):
    """Parse a file, return (session_keys, sign_map, score_map, students, leave_data)."""
    wb = load_workbook(path, data_only=True)
    try:
        return parse_file(wb)
    finally:
        wb.close()


def merge(stus1, stus2):
    """Merge two student lists by ID (order-independent)."""
    stus2_by_id = {s['id']: s for s in stus2}
    return [{
        'id': s1['id'], 'name': s1['name'], 'cls': s1['cls'],
        'attendance': {**s1['attendance'], **stus2_by_id[s1['id']]['attendance']},
        'scores': {**s1['scores'], **stus2_by_id[s1['id']]['scores']},
    } for s1 in stus1]


def run_merge(path1, path2):
    """Simulate merge mode: parse two files, validate, merge, generate output."""
    sk1, ssm1, sscm1, stus1, ld1 = parse(path1)
    sk2, ssm2, sscm2, stus2, ld2 = parse(path2)

    # Validate
    ids1 = [(s['id'], s['name']) for s in stus1]
    ids2 = [(s['id'], s['name']) for s in stus2]
    assert ids1 == ids2, 'Student lists must match'

    # Merge
    sk = sk1 + sk2
    ssm = OrderedDict(list(ssm1.items()) + list(ssm2.items()))
    sscm = OrderedDict(list(sscm1.items()) + list(sscm2.items()))
    stus = merge(stus1, stus2)
    ld = {**ld1, **ld2}
    return sk, ssm, sscm, stus, ld


# ════════════════════════════════════════════
# TEST 1: 单表模式 — 单表
# ════════════════════════════════════════════
section('Test 1: 单表模式 — 单表（单次课导出）')

sk, ssm, sscm, stus, ld = parse(S1)
check_eq(len(sk), 1, '解析到 1 次课')
check_eq(len(stus), 10, '解析到 10 名学生')
# This is 单表测试表1 = 0318实验课
check('实验课' in sk[0] or '实验' in sk[0], f'课程名称: {sk[0]}')

# ════════════════════════════════════════════
# TEST 2: 单表模式 — 汇总表
# ════════════════════════════════════════════
section('Test 2: 单表模式 — 汇总表（完整批量导出）')

sk, ssm, sscm, stus, ld = parse(M1)
check_eq(len(sk), 3, '解析到 3 次课')
check_eq(len(stus), 10, '解析到 10 名学生')
for s in stus[:3]:
    check(len(s['attendance']) == 3, f'{s["name"]} 有 3 次课记录')

# ════════════════════════════════════════════
# TEST 3: 合并 — 单表 + 单表
# ════════════════════════════════════════════
section('Test 3: 合并模式 — 单表 + 单表')

sk, ssm, sscm, stus, ld = run_merge(S1, S2)
check_eq(len(sk), 2, '合并后 2 次课')
check_eq(len(stus), 10, '合并后 10 名学生')
for s in stus[:3]:
    check(len(s['attendance']) == 2, f'{s["name"]} 有 2 次课记录')

# Verify output generation succeeds
buf, info = generate_output(sk, ssm, sscm, stus, ld)
check(isinstance(buf, BytesIO), '输出生成成功')

# ════════════════════════════════════════════
# TEST 4: 合并 — 单表 + 汇总表
# ════════════════════════════════════════════
section('Test 4: 合并模式 — 单表 + 汇总表')

sk, ssm, sscm, stus, ld = run_merge(S1, M1)
check_eq(len(sk), 4, '合并后 4 次课（1 单表 + 3 汇总表）')
check_eq(len(stus), 10, '合并后 10 名学生')

buf, info = generate_output(sk, ssm, sscm, stus, ld)
check(isinstance(buf, BytesIO), '输出生成成功')

# ════════════════════════════════════════════
# TEST 5: 合并 — 汇总表 + 汇总表
# ════════════════════════════════════════════
section('Test 5: 合并模式 — 汇总表 + 汇总表')

sk, ssm, sscm, stus, ld = run_merge(M1, M2)
check_eq(len(sk), 6, '合并后 6 次课（3 + 3）')
check_eq(len(stus), 10, '合并后 10 名学生')

buf, info = generate_output(sk, ssm, sscm, stus, ld)
check(isinstance(buf, BytesIO), '输出生成成功')

# ════════════════════════════════════════════
# TEST 6: 一致性校验
# ════════════════════════════════════════════
section('Test 6: 学生名单不一致检测')

_, _, _, stus1, _ = parse(S1)
_, _, _, stus2, _ = parse(M1)
bad = [dict(s) for s in stus1]
bad[0]['id'] = '99999999'
ids_set1 = {(s['id'], s['name']) for s in bad}
ids_set2 = {(s['id'], s['name']) for s in stus2}
diff = [f'文件一有但文件二缺少：学号 {s[0]} {s[1]}' for s in ids_set1 - ids_set2]
diff += [f'文件二有但文件一缺少：学号 {s[0]} {s[1]}' for s in ids_set2 - ids_set1]
check(len(diff) > 0, f'名单不匹配时检测到 {len(diff)} 处差异')
check(any('99999999' in d for d in diff), '差异中包含修改的学号')

# ════════════════════════════════════════════
# TEST 7: 输出格式验证（考勤明细 + 课堂表现 + 过程性成绩）
# ════════════════════════════════════════════
section('Test 7: 输出格式验证')

sk, ssm, sscm, stus, ld = parse(M1)
buf, info = generate_output(sk, ssm, sscm, stus, ld)

# 考勤明细
out_wb = load_workbook(buf, data_only=False)
ws = out_wb.active
headers = [str(ws.cell(1, c).value or '') for c in range(1, ws.max_column + 1)]
check('无故旷课率' in ''.join(headers[-3:]), '考勤明细有无故旷课率列')
check('总旷课率' in ''.join(headers[-3:]), '考勤明细有总旷课率列')

# 课堂表现
ws_score = out_wb[out_wb.sheetnames[1]]
h_score = [str(ws_score.cell(1, c).value or '') for c in range(1, ws_score.max_column + 1)]
check('课堂表现' in out_wb.sheetnames[1], f'Sheet2 为课堂表现')
out_wb.close()

# 过程性成绩记载表
ps_buf = info.get('process_score_buf')
check(ps_buf is not None, 'info 包含 process_score_buf')
if ps_buf:
    ps_wb = load_workbook(ps_buf, data_only=False)
    ps_ws = ps_wb.active
    ps_headers = [str(ps_ws.cell(1, c).value or '') for c in range(1, ps_ws.max_column + 1)]
    for col in ['序号', '学号', '姓名', '班级']:
        check(col in ps_headers, f'过程性成绩包含列: {col}')
    cell = ps_ws.cell(2, 1)
    check(cell.font.name == 'SimSun', f'字体为宋体, 实际: {cell.font.name}')
    check(cell.font.size == 9, f'字号 9pt, 实际: {cell.font.size}')
    check(cell.alignment.horizontal == 'center', '居中对齐')
    check(cell.border.left.style == 'thin', '细边框')
    ps_wb.close()

# ════════════════════════════════════════════
# SUMMARY
# ════════════════════════════════════════════
print(f'\n{"="*60}')
print(f'  测试完成: {PASS} 通过, {FAIL} 失败')
print(f'{"="*60}')
if FAIL:
    sys.exit(1)
