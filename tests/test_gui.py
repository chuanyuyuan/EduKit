"""
考勤分析工具 (GUI 版) — 功能测试
=================================
测试场景（与 test_app.py 一致）:
  1. 单表模式 — 单表
  2. 单表模式 — 汇总表
  3. 合并模式 — 单表 + 单表
  4. 合并模式 — 单表 + 汇总表
  5. 合并模式 — 汇总表 + 汇总表
  6. 学生名单不一致检测
  7. 输出格式验证
  8. parse_file 自动识别
  9. 输出内容验证
  10. 过程性成绩符号验证

运行： python tests/test_gui.py
"""
import sys, os, glob
sys.stdout.reconfigure(encoding='utf-8')

# ── Import attendance_gui.py ──
PROJECT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEST_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, PROJECT)

import importlib.util
_spec = importlib.util.spec_from_file_location("gui", os.path.join(PROJECT, 'attendance_gui.py'))
gui = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(gui)

from collections import OrderedDict
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border

PASS = FAIL = 0

def check(cond, msg):
    global PASS, FAIL
    if cond: PASS += 1; print(f'  ✓ {msg}')
    else:    FAIL += 1; print(f'  ✗ {msg}')

def check_eq(a, b, msg):
    global PASS, FAIL
    if a == b: PASS += 1; print(f'  ✓ {msg} ({a})')
    else:      FAIL += 1; print(f'  ✗ {msg}: got {a}, expected {b}')

def section(name):
    print(f'\n{"="*60}\n  {name}\n{"="*60}')


# ── Load test files, classify by type ──
_single_files = []
_summary_files = []

for fpath in sorted(glob.glob(os.path.join(TEST_DIR, '*.xlsx'))):
    wb = load_workbook(fpath, data_only=True)
    if len(wb.sheetnames) == 1 and '课堂情况' in wb.sheetnames[0]:
        _single_files.append(fpath)
    elif len(wb.sheetnames) > 1:
        _summary_files.append(fpath)
    wb.close()

assert len(_single_files) >= 2, f'需要 ≥2 个单表文件, 找到 {len(_single_files)}'
assert len(_summary_files) >= 2, f'需要 ≥2 个汇总表文件, 找到 {len(_summary_files)}'

S1 = _single_files[0]
S2 = _single_files[1]
M1 = _summary_files[0]
M2 = _summary_files[1]

print(f'测试文件 (GUI 版):')
print(f'  单表1: {os.path.basename(S1)}')
print(f'  单表2: {os.path.basename(S2)}')
print(f'  汇总表1: {os.path.basename(M1)}')
print(f'  汇总表2: {os.path.basename(M2)}')


def parse(path):
    wb = load_workbook(path, data_only=True)
    try:
        return gui.parse_file(wb)
    finally:
        wb.close()


def merge(stus1, stus2):
    """Merge two student lists by ID (order-independent)."""
    stus2_by_id = {s['id']: s for s in stus2}
    return [{
        'id': s1['id'], 'name': s1['name'], 'cls': s1['cls'],
        'attendance': {**s1['attendance'], **stus2_by_id[s1['id']]['attendance']},
        'scores': {**s1['scores'], **stus2_by_id[s1['id']]['scores']},
    } for s1 in stus1]


def run_merge(path1, path2):
    sk1, ssm1, sscm1, stus1, ld1 = parse(path1)
    sk2, ssm2, sscm2, stus2, ld2 = parse(path2)
    ids1 = [(s['id'], s['name']) for s in stus1]
    ids2 = [(s['id'], s['name']) for s in stus2]
    assert ids1 == ids2, 'Student lists must match'
    sk = sk1 + sk2
    ssm = OrderedDict(list(ssm1.items()) + list(ssm2.items()))
    sscm = OrderedDict(list(sscm1.items()) + list(sscm2.items()))
    stus = merge(stus1, stus2)
    ld = {**ld1, **ld2}
    return sk, ssm, sscm, stus, ld


GET_STATUS = {}  # populated by test 7


# ════════════════════════════════════════════
# TEST 1: 单表模式 — 单表
# ════════════════════════════════════════════
section('Test 1: 单表模式 — 单表')

sk, ssm, sscm, stus, ld = parse(S1)
check_eq(len(sk), 1, '解析到 1 次课')
check_eq(len(stus), 10, '解析到 10 名学生')
for s in stus:
    check(len(s['attendance']) == 1, f'{s["name"]} 有 1 次课记录')

# ════════════════════════════════════════════
# TEST 2: 单表模式 — 汇总表
# ════════════════════════════════════════════
section('Test 2: 单表模式 — 汇总表')

sk, ssm, sscm, stus, ld = parse(M1)
check_eq(len(sk), 3, '解析到 3 次课')
check_eq(len(stus), 10, '解析到 10 名学生')
for s in stus:
    check(len(s['attendance']) == 3, f'{s["name"]} 有 3 次课记录')

# ════════════════════════════════════════════
# TEST 3: 合并 — 单表 + 单表
# ════════════════════════════════════════════
section('Test 3: 合并模式 — 单表 + 单表')

sk, ssm, sscm, stus, ld = run_merge(S1, S2)
check_eq(len(sk), 2, '合并后 2 次课')
check_eq(len(stus), 10, '合并后 10 名学生')
for s in stus:
    check(len(s['attendance']) == 2, f'{s["name"]} 有 2 次课记录')
buf, info = gui.generate_output(sk, ssm, sscm, stus, ld)
check(isinstance(buf, BytesIO), '输出生成成功')

# ════════════════════════════════════════════
# TEST 4: 合并 — 单表 + 汇总表
# ════════════════════════════════════════════
section('Test 4: 合并模式 — 单表 + 汇总表')

sk, ssm, sscm, stus, ld = run_merge(S1, M1)
check_eq(len(sk), 4, '合并后 4 次课')
check_eq(len(stus), 10, '合并后 10 名学生')
buf, info = gui.generate_output(sk, ssm, sscm, stus, ld)
check(isinstance(buf, BytesIO), '输出生成成功')

# ════════════════════════════════════════════
# TEST 5: 合并 — 汇总表 + 汇总表
# ════════════════════════════════════════════
section('Test 5: 合并模式 — 汇总表 + 汇总表')

sk, ssm, sscm, stus, ld = run_merge(M1, M2)
check_eq(len(sk), 6, '合并后 6 次课')
check_eq(len(stus), 10, '合并后 10 名学生')
buf, info = gui.generate_output(sk, ssm, sscm, stus, ld)
check(isinstance(buf, BytesIO), '输出生成成功')

# ════════════════════════════════════════════
# TEST 6: 一致性校验
# ════════════════════════════════════════════
section('Test 6: 学生名单不一致检测')

_, _, _, stus1, _ = parse(S1)
_, _, _, stus2, _ = parse(M1)
bad = [dict(s) for s in stus1]
bad[0]['id'] = '99999999'
s1 = {(s['id'], s['name']) for s in bad}
s2 = {(s['id'], s['name']) for s in stus2}
diff = [f'文件一有但文件二缺少：学号 {s[0]} {s[1]}' for s in s1 - s2]
diff += [f'文件二有但文件一缺少：学号 {s[0]} {s[1]}' for s in s2 - s1]
check(len(diff) > 0, f'名单不匹配时检测到 {len(diff)} 处差异')
check(any('99999999' in d for d in diff), '差异中包含修改的学号')

# ════════════════════════════════════════════
# TEST 7: 输出格式验证
# ════════════════════════════════════════════
section('Test 7: 输出格式验证')

sk, ssm, sscm, stus, ld = parse(M1)
buf, info = gui.generate_output(sk, ssm, sscm, stus, ld)

out_wb = load_workbook(buf, data_only=False)
ws = out_wb.active
headers = [str(ws.cell(1, c).value or '') for c in range(1, ws.max_column + 1)]
check('无故旷课率' in ''.join(headers[-3:]), '考勤明细有无故旷课率列')
check('总旷课率' in ''.join(headers[-3:]), '考勤明细有总旷课率列')
check('课堂表现' in out_wb.sheetnames[1], f'Sheet2 为课堂表现')
out_wb.close()

# GUI 版 generate_output 不返回 process_score_buf，需单独调用
ps_buf = gui.generate_process_score_sheet(stus, sk, ld)
check(isinstance(ps_buf, BytesIO), '过程性成绩表生成成功')
ps_wb = load_workbook(ps_buf, data_only=False)
ps_ws = ps_wb.active
ps_headers = [str(ps_ws.cell(1, c).value or '') for c in range(1, ps_ws.max_column + 1)]
for col in ['序号', '学号', '姓名', '班级']:
    check(col in ps_headers, f'过程性成绩包含列: {col}')
cell = ps_ws.cell(2, 1)
check(cell.font.name == 'SimSun', f'字体为宋体, 实际: {cell.font.name}')
check(cell.font.size == 9, f'字号 9pt, 实际: {cell.font.size}')
check(cell.alignment.horizontal == 'center', '居中对齐')
check(cell.border.left.style == 'thin', '细边框')
ps_wb.close()

# ════════════════════════════════════════════
# TEST 8: parse_file 自动识别
# ════════════════════════════════════════════
section('Test 8: parse_file 自动识别')

# 单表文件通过 parse_single_session 解析
sk_s, ssm_s, sscm_s, stus_s, ld_s = parse(S1)
check_eq(len(sk_s), 1, '单表文件识别为 1 次课')
SESSION_NAME = sk_s[0]  # save for later

# 汇总表文件通过 parse_summary + parse_sub_sheets 解析
sk_m, ssm_m, sscm_m, stus_m, ld_m = parse(M1)
check_eq(len(sk_m), 3, '汇总表文件识别为 3 次课')
check(stus_s[0]['id'] == stus_m[0]['id'], '两文件学号一致')

# col_idx 工具函数
check_eq(gui.col_idx('A'), 0, 'col_idx(A) = 0')
check_eq(gui.col_idx('Z'), 25, 'col_idx(Z) = 25')
check_eq(gui.col_idx('AA'), 26, 'col_idx(AA) = 26')
check_eq(gui.col_idx('AB'), 27, 'col_idx(AB) = 27')

# ════════════════════════════════════════════
# TEST 9: 输出内容验证
# ════════════════════════════════════════════
section('Test 9: 输出内容验证')

sk, ssm, sscm, stus, ld = parse(M1)
buf, info = gui.generate_output(sk, ssm, sscm, stus, ld)
out_wb = load_workbook(buf, data_only=False)
ws = out_wb.active

# 验证数据行数 = 10 学生 + 1 表头
check_eq(ws.max_row, 11, '考勤明细行数 = 11（含表头）')
check_eq(ws.max_column, 2 + 3 + 2, '考勤明细列数 = 7（学号+姓名+3次课+2比率）')

# 统计每种考勤状态的出现次数
from collections import Counter
status_cnt = Counter()
for r in range(2, ws.max_row + 1):
    for c in range(3, 3 + 3):  # attendance columns
        v = str(ws.cell(r, c).value or '')
        status_cnt[v] += 1
check(status_cnt.get('上课', 0) >= 20, f'上课状态合计 ≥20 次 ({status_cnt.get("上课", 0)})')
check(status_cnt.get('旷课', 0) >= 3, f'旷课状态合计 ≥3 次 ({status_cnt.get("旷课", 0)})')
total = sum(status_cnt.values())
check_eq(total, 30, '考勤状态总数 = 30（10学生 × 3次课）')

# 验证颜色填充
green = '00C6EFCE'
red = '00FFC7CE'
yellow_fills = ['00FFEB9C']
fill_ok = 0
for r in range(2, ws.max_row + 1):
    for c in range(3, 3 + 3):
        cell = ws.cell(r, c)
        if cell.value == '上课' and cell.fill.start_color.rgb == green:
            fill_ok += 1
        elif cell.value == '旷课' and cell.fill.start_color.rgb == red:
            fill_ok += 1
check(fill_ok >= 20, f'颜色标注正确 ≥20 个单元格 ({fill_ok})')

# 验证比率列格式
rate_col = 3 + 3  # first ratio column
for r in range(2, ws.max_row + 1):
    v = str(ws.cell(r, rate_col).value or '')
    check(v.endswith('%'), f'第 {r} 行旷课率格式: {v}')
out_wb.close()

# ════════════════════════════════════════════
# TEST 10: 过程性成绩符号验证
# ════════════════════════════════════════════
section('Test 10: 过程性成绩符号验证')

ps_buf = gui.generate_process_score_sheet(stus, sk, ld)
ps_wb = load_workbook(ps_buf, data_only=False)
ps_ws = ps_wb.active

# 行数 = 10 学生 + 1 表头
check_eq(ps_ws.max_row, 11, '过程性成绩表行数 = 11')
# 列：序号 学号 姓名 性别 专业 班级 + 3次课 + 填充 + 成绩 备注
check_eq(ps_ws.max_column >= 10, True, f'过程性成绩表列数 ≥ 10 ({ps_ws.max_column})')

# 验证符号：只能出现 ✓ ✗ △
valid_symbols = {'✓', '✗', '△'}
symbol_cnt = Counter()
for r in range(2, ps_ws.max_row + 1):
    for c in range(7, 7 + 3):  # session columns (G, H, I)
        v = str(ps_ws.cell(r, c).value or '')
        if v in valid_symbols:
            symbol_cnt[v] += 1
total_symbols = sum(symbol_cnt.values())
check_eq(total_symbols, 30, f'符号总数 = 30 ({dict(symbol_cnt)})')
check(symbol_cnt.get('✓', 0) >= 20, f'✓ 出现 ≥20 次 ({symbol_cnt.get("✓", 0)})')
check(symbol_cnt.get('✗', 0) >= 3, f'✗ 出现 ≥3 次 ({symbol_cnt.get("✗", 0)})')

# 验证每个学生行有正确的列数
for r in range(2, ps_ws.max_row + 1):
    sid = str(ps_ws.cell(r, 2).value or '')
    name = str(ps_ws.cell(r, 3).value or '')
    check(sid.startswith('0423'), f'{name} 学号以 0423 开头: {sid}')

ps_wb.close()

# ════════════════════════════════════════════
# SUMMARY
# ════════════════════════════════════════════
section('SUMMARY')
print(f'  GUI 版: {PASS} 通过, {FAIL} 失败')
print(f'{"="*60}')
if FAIL:
    sys.exit(1)
