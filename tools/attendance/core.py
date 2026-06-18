"""
雨课堂考勤分析 — 核心逻辑
解析、合并、生成输出文件，无 UI 依赖。
"""
from collections import OrderedDict
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 常量 ──

PRESENT_SET = {'扫二维码', '“正在上课”提示', '教师添加', '课堂暗号'}


# ── 工具函数 ──

def col_idx(col):
    i = 0
    for ch in col:
        i = i * 26 + (ord(ch) - 64)
    return i - 1


# ── 解析函数 ──

def parse_summary(wb):
    if not wb.sheetnames:
        raise ValueError('Excel 文件中没有找到任何工作表。')

    ws = wb[wb.sheetnames[0]]

    row2 = {}
    for cell in ws[2]:
        if cell.value:
            row2[cell.column_letter] = str(cell.value)

    if not row2:
        raise ValueError('汇总表第 2 行为空，未找到子表头信息。请确认上传的是雨课堂批量导出文件（含汇总页）。')

    row1 = {}
    for cell in ws[1]:
        if cell.value:
            row1[cell.column_letter] = str(cell.value)

    if not row1:
        raise ValueError('汇总表第 1 行为空，未找到课堂名称合并表头。')

    sign_cols = sorted(
        [c for c, v in row2.items() if v == '签到方式'],
        key=col_idx
    )
    if not sign_cols:
        raise ValueError('未找到"签到方式"列。请确认上传的是雨课堂导出的"汇总-数据表"文件。')

    score_cols = sorted(
        [c for c, v in row2.items() if v.startswith('得分')],
        key=col_idx
    )
    session_headers = sorted(row1.keys(), key=col_idx)

    def map_cols_to_sessions(cols):
        result = OrderedDict()
        keys = []
        for c in cols:
            ci = col_idx(c)
            nearest = None
            nearest_i = -1
            for sh in session_headers:
                hi = col_idx(sh)
                if hi <= ci and hi > nearest_i:
                    nearest = row1[sh]
                    nearest_i = hi
            if nearest:
                result[nearest] = c
                keys.append(nearest)
        return result, keys

    session_sign_map, session_keys = map_cols_to_sessions(sign_cols)
    session_score_map, _ = map_cols_to_sessions(score_cols)

    students = []
    for row in ws.iter_rows(min_row=3, values_only=False):
        cells = {}
        for c in row:
            if c.value is None:
                continue
            try:
                cells[c.column_letter] = str(c.value)
            except AttributeError:
                pass
        sid = cells.get('A', '')
        name = cells.get('D', '')
        if not sid and not name:
            continue

        rec = {
            'id': sid,
            'name': name,
            'dept': cells.get('B', ''),
            'cls': cells.get('C', ''),
            'attendance': {},
            'scores': {},
        }
        for sk, sc in session_sign_map.items():
            rec['attendance'][sk] = cells.get(sc, '')
        for sk, sc in session_score_map.items():
            raw = cells.get(sc, '')
            try:
                rec['scores'][sk] = int(float(raw)) if raw else None
            except ValueError:
                rec['scores'][sk] = None
        students.append(rec)

    return session_keys, session_sign_map, session_score_map, students


def parse_sub_sheets(wb, session_keys):
    leave_data = {}
    idx = 0
    found = False

    for name in wb.sheetnames:
        if '课堂情况' not in name:
            continue
        found = True
        if idx >= len(session_keys):
            break

        session_key = session_keys[idx]
        idx += 1
        ws = wb[name]

        for row in ws.iter_rows(min_row=4, values_only=False):
            cells = {}
            for c in row:
                if c.value is None:
                    continue
                try:
                    cells[c.column_letter] = str(c.value)
                except AttributeError:
                    pass
            sid = cells.get('A', '')
            remark = cells.get('G', '')
            if sid and remark in ('病假', '事假'):
                leave_data[(session_key, sid)] = remark

    if not found:
        raise ValueError('未找到名称包含"课堂情况"的子表。请确认上传的是雨课堂批量导出文件（含每次课的课堂情况子表）。')

    return leave_data


def parse_single_session(wb):
    """Parse a single-session export file (no summary page, just one 课堂情况 sheet)."""
    ws = wb[wb.sheetnames[0]]

    raw_name = str(ws.cell(1, 1).value or '').strip()
    if not raw_name:
        raise ValueError('文件第 1 行为空，未找到课堂名称。')

    row3_col5 = str(ws.cell(3, 5).value or '')
    if '签到方式' not in row3_col5:
        raise ValueError('第 3 行未找到"签到方式"列，请确认是雨课堂导出的课堂数据表。')

    students = []
    for row_idx in range(5, ws.max_row + 1):
        sid = str(ws.cell(row_idx, 1).value or '').strip()
        name = str(ws.cell(row_idx, 4).value or '').strip()
        if not sid and not name:
            continue
        students.append({
            'id': sid,
            'name': name,
            'dept': '',
            'cls': '',
            'attendance': {},
            'scores': {},
        })
        students[-1]['_sign'] = str(ws.cell(row_idx, 5).value or '')
        students[-1]['_remark'] = str(ws.cell(row_idx, 7).value or '')

    session_keys = [raw_name]
    session_sign_map = OrderedDict([(raw_name, raw_name)])
    session_score_map = OrderedDict()

    leave_data = {}
    for s in students:
        s['attendance'][raw_name] = s.pop('_sign')
        remark = s.pop('_remark')
        if remark in ('病假', '事假'):
            leave_data[(raw_name, s['id'])] = remark

    return session_keys, session_sign_map, session_score_map, students, leave_data


def parse_file(wb):
    """Auto-detect file type (full export or single session) and parse."""
    if len(wb.sheetnames) == 1 and '课堂情况' in wb.sheetnames[0]:
        return parse_single_session(wb)
    sk, ssm, sscm, students = parse_summary(wb)
    ld = parse_sub_sheets(wb, sk)
    return sk, ssm, sscm, students, ld


# ── 输出生成 ──

def generate_output(session_keys, session_sign_map, session_score_map, students, leave_data):
    """Generate output Excel bytes + summary info from parsed data."""

    session_count = len(session_sign_map)

    out_rows = []
    for s in students:
        row = [s['id'], s['name']]
        absent_cnt = 0
        excused_cnt = 0
        for sk in session_sign_map:
            raw = s['attendance'].get(sk, '')
            if raw in PRESENT_SET:
                status = '上课'
            elif raw == '未上课':
                leave = leave_data.get((sk, s['id']))
                if leave == '病假':
                    status = '病假'
                    excused_cnt += 1
                elif leave == '事假':
                    status = '事假'
                    excused_cnt += 1
                else:
                    status = '旷课'
                    absent_cnt += 1
            else:
                status = raw
            row.append(status)
        row.append(f'{absent_cnt / session_count:.0%}')
        row.append(f'{(absent_cnt + excused_cnt) / session_count:.0%}')
        out_rows.append(row)

    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    color_map = {'上课': green_fill, '病假': yellow_fill, '事假': yellow_fill, '旷课': red_fill}
    gold_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')

    owb = Workbook()
    ows = owb.active
    ows.title = '考勤明细'
    headers = ['学号', '姓名'] + list(session_sign_map.keys()) + ['无故旷课率', '总旷课率']
    ows.append(headers)
    for row in out_rows:
        ows.append(row)
    for row in ows.iter_rows(min_row=2, max_row=ows.max_row,
                             min_col=3, max_col=2 + session_count):
        for cell in row:
            fill = color_map.get(cell.value)
            if fill:
                cell.fill = fill

    ows2 = owb.create_sheet('课堂表现')
    score_headers = ['学号', '姓名'] + [k for k in session_score_map.keys()] + ['总分']
    ows2.append(score_headers)
    totals = []
    for s in students:
        row = [s['id'], s['name']]
        total = 0
        for sk in session_score_map:
            v = s['scores'].get(sk)
            if v is not None:
                row.append(v)
                total += v
            else:
                row.append('')
        row.append(total)
        totals.append(total)
        ows2.append(row)

    sorted_totals = sorted(totals, reverse=True)
    n_top = max(1, round(len(sorted_totals) * 0.1))
    threshold = sorted_totals[n_top - 1] if n_top <= len(sorted_totals) else sorted_totals[-1]
    for row in ows2.iter_rows(min_row=2, max_row=ows2.max_row):
        if row[-1].value is not None and row[-1].value >= threshold:
            for cell in row:
                cell.fill = gold_fill

    # 每次课统计摘要
    summary_lines = []
    for sk in session_sign_map:
        attend = absent = sick = personal = 0
        for s in students:
            raw = s['attendance'].get(sk, '')
            if raw in PRESENT_SET:
                attend += 1
            elif raw == '未上课':
                leave = leave_data.get((sk, s['id']))
                if leave == '病假':
                    sick += 1
                elif leave == '事假':
                    personal += 1
                else:
                    absent += 1
        summary_lines.append({
            'session': sk,
            '上课': attend,
            '旷课': absent,
            '病假': sick,
            '事假': personal,
        })

    ows3 = owb.create_sheet('考勤统计')
    ows3.append(['课程', '上课', '旷课', '病假', '事假'])
    for cell in ows3[1]:
        cell.font = Font(bold=True)
    for line in summary_lines:
        ows3.append([line['session'], line['上课'], line['旷课'], line['病假'], line['事假']])

    buf = BytesIO()
    owb.save(buf)
    buf.seek(0)

    total_absent = sum(1 for s in students for sk in session_sign_map
                       if s['attendance'].get(sk) == '未上课'
                       and (sk, s['id']) not in leave_data)
    total_sick = sum(1 for v in leave_data.values() if v == '病假')
    total_personal = sum(1 for v in leave_data.values() if v == '事假')

    session_names = list(session_sign_map.keys())
    preview_attendance = []
    for s in students:
        row = {'学号': s['id'], '姓名': s['name']}
        absent_cnt = 0
        excused_cnt = 0
        for sk in session_names:
            raw = s['attendance'].get(sk, '')
            if raw in PRESENT_SET:
                status = '上课'
            elif raw == '未上课':
                leave = leave_data.get((sk, s['id']))
                if leave == '病假':
                    status = '病假'
                    excused_cnt += 1
                elif leave == '事假':
                    status = '事假'
                    excused_cnt += 1
                else:
                    status = '旷课'
                    absent_cnt += 1
            else:
                status = raw
            row[sk] = status
        row['无故旷课率'] = f'{absent_cnt / session_count:.0%}'
        row['总旷课率'] = f'{(absent_cnt + excused_cnt) / session_count:.0%}'
        preview_attendance.append(row)

    score_names = list(session_score_map.keys())
    preview_scores = []
    for s in students:
        row = {'学号': s['id'], '姓名': s['name']}
        total = 0
        for k in score_names:
            v = s['scores'].get(k)
            if v is not None:
                row[k] = v
                total += v
            else:
                row[k] = ''
        row['总分'] = total
        preview_scores.append(row)

    return buf, {
        'students': students,
        'session_count': session_count,
        'total_absent': total_absent,
        'total_sick': total_sick,
        'total_personal': total_personal,
        'summary_lines': summary_lines,
        'preview_headers': ['学号', '姓名'] + session_names + ['无故旷课率', '总旷课率'],
        'preview_attendance': preview_attendance,
        'score_headers': ['学号', '姓名'] + score_names + ['总分'],
        'preview_scores': preview_scores,
        'session_keys': session_keys,
        'leave_data': leave_data,
        'process_score_buf': generate_process_score_sheet(students, session_keys, leave_data),
    }


def generate_process_score_sheet(students, session_keys, leave_data):
    """生成过程性成绩记载表（✓/✗/△），与样表格式一致"""
    owb = Workbook()
    ows = owb.active
    ows.title = '过程性成绩记载表'

    n_sessions = len(session_keys)
    session_start = 7
    score_col = max(session_start + n_sessions, 26)
    remark_col = score_col + 1

    headers = ['序号', '学号', '姓名', '性别', '专业', '班级'] + list(session_keys)
    while len(headers) < score_col - 1:
        headers.append('')
    headers.append('成绩')
    headers.append('备注')
    ows.append(headers)

    for i, s in enumerate(students, start=1):
        row = [i, s['id'], s['name'], '', '', s.get('cls', '')]
        for sk in session_keys:
            raw = s['attendance'].get(sk, '')
            if raw in PRESENT_SET:
                row.append('✓')
            elif raw == '未上课':
                leave = leave_data.get((sk, s['id']))
                row.append('△' if leave in ('病假', '事假') else '✗')
            else:
                row.append(raw)
        while len(row) < score_col - 1:
            row.append('')
        row.append('')
        row.append('')
        ows.append(row)

    font_song = Font(name='SimSun', size=9)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for row in ows.iter_rows(min_row=1, max_row=ows.max_row,
                              min_col=1, max_col=remark_col):
        for cell in row:
            cell.font = font_song
            cell.alignment = align_center
            cell.border = thin_border

    col_widths = {
        'A': 2.85, 'B': 12.75, 'C': 11.26, 'D': 3.25,
        'E': 15.87, 'F': 15.47,
    }
    for col_letter, width in col_widths.items():
        ows.column_dimensions[col_letter].width = width
    for i in range(session_start, score_col):
        ows.column_dimensions[get_column_letter(i)].width = 3.69
    ows.column_dimensions[get_column_letter(score_col)].width = 5.38
    ows.column_dimensions[get_column_letter(remark_col)].width = 9.77

    ows.row_dimensions[1].height = 22.6
    for r in range(2, ows.max_row + 1):
        ows.row_dimensions[r].height = 14.3

    buf = BytesIO()
    owb.save(buf)
    buf.seek(0)
    return buf


def merge_datasets(datasets: list) -> tuple:
    """Merge multiple parsed datasets into one.

    datasets: list of (session_keys, session_sign_map, session_score_map, students, leave_data)
    Returns: (session_keys, session_sign_map, session_score_map, students, leave_data)

    All datasets must have the same student roster (same IDs + names).
    Sessions are concatenated in order; attendance/scores per student are merged by ID.
    """
    if not datasets:
        raise ValueError("没有数据可合并")

    session_keys = []
    session_sign_map = OrderedDict()
    session_score_map = OrderedDict()
    leave_data = {}

    for sk, ssm, sscm, students, ld in datasets:
        session_keys.extend(sk)
        session_sign_map.update(ssm)
        session_score_map.update(sscm)
        leave_data.update(ld)

    # Merge each student's attendance and scores by ID across all files
    students_by_id = {}
    for _, _, _, students, _ in datasets:
        for s in students:
            sid = s['id']
            if sid not in students_by_id:
                students_by_id[sid] = {
                    k: v for k, v in s.items() if k not in ('attendance', 'scores')
                }
                students_by_id[sid]['attendance'] = {}
                students_by_id[sid]['scores'] = {}
            students_by_id[sid]['attendance'].update(s['attendance'])
            students_by_id[sid]['scores'].update(s['scores'])

    merged_students = list(students_by_id.values())

    return session_keys, session_sign_map, session_score_map, merged_students, leave_data
