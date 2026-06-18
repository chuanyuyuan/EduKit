"""
雨课堂考勤分析 — Streamlit UI 组件
"""
import os
import base64
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

from .core import parse_file, generate_output, generate_process_score_sheet, merge_datasets

DEMO_FILE = 'samples/sample_attendance.xlsx'


def _sample_link():
    """Return base64 data URI for the demo file download link."""
    path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), DEMO_FILE)
    if not os.path.exists(path):
        return ''
    with open(path, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode()
    return f'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}'


def show_results(buf, info):
    """Display preview tables and download buttons."""
    st.subheader(f"考勤概况 — 共 {info['session_count']} 次课，{len(info['students'])} 名学生")

    tab1, tab2, tab3 = st.tabs([":material/table: 考勤明细", ":material/star: 课堂表现", ":material/bar_chart: 统计摘要"])

    with tab1:
        df = pd.DataFrame(info['preview_attendance'])
        if df.empty:
            st.info("暂无考勤数据。")
            return
        styled = df.style.map(
            lambda v: 'background-color: #C6EFCE' if v == '上课'
            else 'background-color: #FFEB9C' if v in ('病假', '事假')
            else 'background-color: #FFC7CE' if v == '旷课'
            else ''
        )
        st.dataframe(styled, use_container_width=True)

    with tab2:
        all_scores = info['preview_scores']
        totals = [s['总分'] for s in all_scores]
        sorted_totals = sorted(totals, reverse=True)
        n_top = max(1, round(len(sorted_totals) * 0.1))
        threshold = sorted_totals[n_top - 1]

        df_scores = pd.DataFrame(all_scores)
        styled_scores = df_scores.style.apply(
            lambda row: ['background-color: #FFD700'] * len(row)
            if row['总分'] >= threshold else [''] * len(row),
            axis=1,
        )
        st.dataframe(styled_scores, use_container_width=True)

    with tab3:
        col1, col2, col3 = st.columns(3)
        col1.metric("旷课总人次", info['total_absent'])
        col2.metric("病假总人次", info['total_sick'])
        col3.metric("事假总人次", info['total_personal'])

        st.subheader(":material/calendar_view_week: 每次课统计")
        summary_df = []
        for line in info['summary_lines']:
            summary_df.append({
                '课程': line['session'],
                '上课': line['上课'],
                '旷课': line['旷课'],
                '病假': line['病假'],
                '事假': line['事假'],
            })
        st.dataframe(summary_df, use_container_width=True, hide_index=True)

    st.divider()
    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        st.download_button(
            label=":material/download: 下载考勤明细 Excel",
            data=buf,
            file_name="学生考勤明细表.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    with col_dl2:
        st.download_button(
            label=":material/description: 下载过程性成绩记载表",
            data=info['process_score_buf'],
            file_name="过程性成绩记载表.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


def single_page():
    """单文件模式页面"""
    uploaded = st.file_uploader("选择雨课堂导出的 Excel 文件（仅支持 .xlsx 格式）")

    workbook_source = None
    source_label = None
    if uploaded:
        if not uploaded.name.endswith('.xlsx'):
            st.error(f'不支持的文件格式："{uploaded.name}"，请上传 .xlsx 文件。')
        else:
            try:
                workbook_source = load_workbook(uploaded, data_only=True)
                source_label = "upload"
                st.session_state.att_show_demo = False
            except Exception:
                st.error("无法读取该文件，请确认上传的是有效的 .xlsx 格式文件。")
    elif st.session_state.get("att_show_demo"):
        demo_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), DEMO_FILE)
        if os.path.exists(demo_path):
            try:
                workbook_source = load_workbook(demo_path, data_only=True)
                source_label = "demo"
            except Exception:
                st.error("无法读取示例文件。")
                st.session_state.att_show_demo = False

    if not uploaded and not st.session_state.get("att_show_demo"):
        if st.button(":material/lightbulb: 加载示例数据看看效果", use_container_width=True):
            st.session_state.att_show_demo = True
            st.rerun()

    if workbook_source:
        if source_label == "demo":
            st.info("当前展示的是示例数据，你可以上传自己的文件替换。")

        try:
            status = st.status("正在解析...", expanded=True)
            with status:
                st.write("读取 Excel 文件...")
                wb = workbook_source
                session_keys, session_sign_map, session_score_map, students, leave_data = parse_file(wb)
                wb.close()
                st.write("生成输出文件...")
                buf, info = generate_output(session_keys, session_sign_map, session_score_map, students, leave_data)
            status.update(label="解析完成", state="complete", expanded=False)
            show_results(buf, info)
        except ValueError as e:
            st.error(str(e))
        except Exception:
            st.error("解析过程出现未知错误，请确认上传的是雨课堂批量导出的汇总数据表（.xlsx）。")
    else:
        st.info("请上传文件开始分析。")


def merge_page():
    """多文件合并模式 — 动态添加上传，逐个校验名单一致性。"""
    MAX_FILES = 20
    MAX_TOTAL_BYTES = 200 * 1024 * 1024

    if "merge_entries" not in st.session_state:
        st.session_state.merge_entries = []
    if "merge_counter" not in st.session_state:
        st.session_state.merge_counter = 0

    entries = st.session_state.merge_entries
    to_remove = None
    base_student_set = None  # 以第一个有效文件为基准

    # ── 获取基准名单 ──
    for entry in entries:
        if entry.get("data"):
            base_entry_students = entry["data"][3]
            base_student_set = {(s["id"], s["name"]) for s in base_entry_students}
            break

    # ── 逐行展示每个文件 ──
    for i, entry in enumerate(entries):
        uploaded = st.file_uploader(
            f"文件 {i + 1}",
            type="xlsx",
            key=f"merge_upload_{entry['key']}",
            label_visibility="collapsed",
        )

        if uploaded:
            file_id = (uploaded.name, uploaded.size)
            if entry.get("file_id") != file_id:
                entry["file_id"] = file_id
                entry["name"] = uploaded.name
                entry["size"] = uploaded.size
                try:
                    wb = load_workbook(uploaded, data_only=True)
                    data = parse_file(wb)
                    wb.close()

                    students = data[3]
                    student_set = {(s["id"], s["name"]) for s in students}
                    if base_student_set is not None and student_set != base_student_set:
                        diff = _roster_diff(base_student_set, student_set)
                        raise ValueError(f"学生名单不一致：\n" + "\n".join(diff))

                    entry["data"] = data
                    entry["status"] = "valid"
                    entry["error"] = None
                except ValueError as e:
                    entry["status"] = "error"
                    entry["error"] = str(e)
                    entry["data"] = None
                except Exception:
                    entry["status"] = "error"
                    entry["error"] = "无法读取该文件，请确认上传的是有效的 .xlsx 格式文件。"
                    entry["data"] = None
                st.rerun()

        # 信息行：文件名 + 状态 + 删除
        if entry.get("name"):
            info_cols = st.columns([3, 1, 0.5, 0.5])
            size_kb = entry["size"] / 1024
            size_str = f"{size_kb:.0f} KB" if size_kb < 1024 else f"{size_kb / 1024:.1f} MB"

            with info_cols[0]:
                st.markdown(f"📄 **{entry['name']}**  — {size_str}")
            with info_cols[1]:
                s = entry.get("status", "empty")
                if s == "valid":
                    st.markdown("✅ **通过**")
                elif s == "error":
                    st.markdown("❌ **不通过**")
            with info_cols[3]:
                if st.button("🗑", key=f"merge_del_{entry['key']}", help="移除此文件"):
                    to_remove = i

        if entry.get("error"):
            st.error(entry["error"])

    if to_remove is not None:
        entries.pop(to_remove)
        st.rerun()

    # ── 添加文件按钮 ──
    if len(entries) < MAX_FILES:
        if st.button("+ 添加文件", use_container_width=True):
            st.session_state.merge_counter += 1
            entries.append({
                "key": st.session_state.merge_counter,
                "file_id": None,
                "name": None,
                "size": 0,
                "status": "empty",
                "error": None,
                "data": None,
            })
            st.rerun()

    # ── 总计信息 ──
    valid_count = sum(1 for e in entries if e.get("status") == "valid")
    total_size = sum(e.get("size", 0) for e in entries if e.get("name"))
    total_mb = total_size / 1024 / 1024
    over_size = total_size > MAX_TOTAL_BYTES

    if entries:
        color = "red" if over_size else "gray"
        st.caption(
            f'{valid_count}/{len(entries)} 个文件通过校验 | '
            f'总计 <span style="color:{color};">{total_mb:.1f} MB</span>（上限 200 MB）',
            unsafe_allow_html=True,
        )

    if over_size:
        st.error("文件总大小超过 200 MB 限制，请移除部分文件。")

    # ── 合并按钮 ──
    can_merge = (
        valid_count >= 2
        and all(e.get("status") == "valid" for e in entries)
        and not over_size
    )
    if st.button("🚀 开始合并", type="primary", use_container_width=True, disabled=not can_merge):
        datasets = [e["data"] for e in entries if e.get("status") == "valid"]
        with st.status("正在合并...", expanded=True) as status:
            st.write("合并数据...")
            sk, ssm, sscm, students, ld = merge_datasets(datasets)
            st.write("生成输出文件...")
            buf, info = generate_output(sk, ssm, sscm, students, ld)
        status.update(label="合并完成", state="complete", expanded=False)
        show_results(buf, info)

    if not entries:
        st.info("点击「添加文件」上传至少 2 份雨课堂文件（单次课或批量导出均可）。")


def _roster_diff(base_set, file_set):
    """比较两个学生集合，返回差异描述列表。"""
    diff = []
    for sid, sname in base_set - file_set:
        diff.append(f"  基准有但该文件缺少：学号 {sid} {sname}")
    for sid, sname in file_set - base_set:
        diff.append(f"  该文件有但基准缺少：学号 {sid} {sname}")
    return diff


def render_attendance_page():
    """考勤分析工具入口 — 说明 + 模式选择 + 单文件/合并模式路由。"""
    st.header(":material/calendar_month: 雨课堂课堂数据分析")
    st.markdown(f"""
上传雨课堂导出的 Excel 文件，自动生成考勤明细和课堂表现统计。

**适用平台：** [长江雨课堂](https://changjiang.yuketang.cn/web/?index)

**支持的文件类型：**
- 批量导出的汇总数据表（含汇总页 + 课堂情况子表）
- 单次课导出的课堂数据表（自动识别）

**原始文件获取方式：**
1. 登录长江雨课堂官网，进入你授课的班级
2. 点击 **批量导出数据**
3. 筛选需要统计的课堂记录
4. 下载导出的 Excel 文件（即 `.xlsx` 格式的汇总数据表）。<a href="{_sample_link()}" download="{DEMO_FILE}">下载示例表格</a>

**上传后：**
- 自动解析并分 tab 展示考勤明细、课堂表现得分和统计摘要
- 支持单文件分析和多文件合并（任意组合单次课/批量导出文件）
- 提供两种下载：考勤明细 Excel（颜色标注）和过程性成绩记载表（✓/✗/△）
""", unsafe_allow_html=True)

    st.markdown("""
    <style>
        button[kind='segmented_control'], button[kind='segmented_controlActive'] {
            font-size: 1.8rem !important;
            padding: 1rem 2.5rem !important;
        }
    </style>
    """, unsafe_allow_html=True)

    mode = st.segmented_control("模式", [":material/file_present: 单文件模式", ":material/merge: 合并模式"],
                                default=":material/file_present: 单文件模式", label_visibility="collapsed")

    if "单文件模式" in mode:
        single_page()
    else:
        merge_page()
