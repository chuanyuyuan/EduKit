"""答辩顺序生成器 — 核心逻辑"""
import random
import io
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def parse_names(text: str) -> list[tuple[str, str]]:
    """从文本中提取学生姓名列表，返回 [(学号, 姓名)]，学号为空字符串。"""
    import re
    names = re.split(r"[\n,，、；;\s]+", text.strip())
    return [("", n.strip()) for n in names if n.strip()]


def parse_roster(file) -> list[tuple[str, str]]:
    """从 Excel 花名册中读取学生列表，返回 [(学号, 姓名)]。

    自动检测包含「学号」和「姓名」的表头列，找不到则抛 ValueError。
    """
    import openpyxl
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # 定位列
    sid_col = name_col = None
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is None:
            continue
        h = str(h).strip()
        if "学号" in h:
            sid_col = c
        elif "姓名" in h:
            name_col = c

    if sid_col is None or name_col is None:
        raise ValueError(
            "未找到「学号」和/或「姓名」列。"
            "请确保表头行包含「学号」和「姓名」字段。"
        )

    students = []
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, sid_col).value
        name = ws.cell(r, name_col).value
        if not name:
            continue
        students.append((str(sid or ""), str(name).strip()))
    wb.close()
    return students


def _assign_times(student_list, start_time, end_time):
    """为给定的 (学号, 姓名) 列表分配时间，返回 (rows, per_student_seconds)。

    rows: [(序号, 学号, 姓名, 开始_str, 结束_str), ...]
    """
    t1 = datetime.strptime(start_time, "%H:%M")
    t2 = datetime.strptime(end_time, "%H:%M")
    if t2 <= t1:
        raise ValueError("结束时间必须晚于开始时间")
    total_seconds = int((t2 - t1).total_seconds())
    per_student = total_seconds / len(student_list) if student_list else 0

    rows = []
    for i, (sid, name) in enumerate(student_list):
        s = t1 + timedelta(seconds=int(i * per_student))
        e = t1 + timedelta(seconds=int((i + 1) * per_student))
        rows.append((i + 1, sid, name, s.strftime("%H:%M"), e.strftime("%H:%M")))

    return rows, per_student


def build_schedule(order: list, start_time: str, end_time: str):
    """按给定顺序和时间区间分配每人时间，不随机打乱。

    order: list of (学号, 姓名) 或 list of str（仅姓名）。
    返回 (rows, per_student_seconds):
      rows: [(序号, 学号, 姓名, 开始_str, 结束_str), ...]
    """
    student_list = [item if isinstance(item, (list, tuple)) else ("", item) for item in order]
    return _assign_times(student_list, start_time, end_time)


def format_duration(seconds: float) -> str:
    m = int(seconds // 60)
    s = round(seconds % 60)
    if s == 0:
        return f"{m}分钟"
    return f"{m}分{s}秒"


def to_excel(rows: list, per_student_seconds: float) -> bytes:
    """生成 Excel 文件并返回二进制内容。

    rows 中的元素可以是:
      - 4 元素: (序号, 姓名, 开始, 结束) — 无学号，旧格式
      - 5 元素: (序号, 学号, 姓名, 开始, 结束) — 含学号
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "答辩顺序表"

    header_font = Font(name="微软雅黑", bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    cell_font = Font(name="微软雅黑", size=10.5)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 判断是否含学号（至少一行有非空学号）
    has_id = any(len(r) == 5 and r[1] for r in rows)

    # 表头
    headers = ["答辩序号", "学号", "姓名", "开始时间", "结束时间"] if has_id else ["答辩序号", "姓名", "开始时间", "结束时间"]
    widths = [12, 16, 16, 14, 14] if has_id else [12, 18, 14, 14]
    for ci, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    # 数据行
    for ri, row in enumerate(rows, start=2):
        seq, sid, name, t1, t2 = row
        vals = [seq, sid, name, t1, t2] if has_id else [seq, name, t1, t2]
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = cell_font
            cell.alignment = center
            cell.border = thin_border

    # 信息行
    duration_str = format_duration(per_student_seconds)
    info_row = len(rows) + 3
    ws.cell(row=info_row, column=1, value="⏱ 每人答辩时间：").font = Font(name="微软雅黑", bold=True, size=10.5)
    ws.cell(row=info_row, column=2, value=duration_str).font = Font(name="微软雅黑", size=10.5)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
