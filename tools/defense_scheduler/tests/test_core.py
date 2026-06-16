"""
答辩顺序生成器 — 核心逻辑测试
"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', '..'))

from tools.defense_scheduler.core import (
    parse_names, parse_roster, build_schedule, format_duration, to_excel,
)
import io
from openpyxl import load_workbook

PASS = FAIL = 0


def check(cond, msg):
    global PASS, FAIL
    if cond: PASS += 1; print(f'  ✓ {msg}')
    else: FAIL += 1; print(f'  ✗ {msg}')


def check_eq(a, b, msg):
    global PASS, FAIL
    if a == b: PASS += 1; print(f'  ✓ {msg} ({a})')
    else: FAIL += 1; print(f'  ✗ {msg}: got {a}, expected {b}')


def section(name):
    print(f'\n{"="*60}\n  {name}\n{"="*60}')


# ════════════════════════════════════════════
# Test 1: parse_names — 返回 [(学号, 姓名)]
# ════════════════════════════════════════════
section('Test 1: parse_names')

check_eq(parse_names(''), [], '空字符串')
check_eq(parse_names('  \n  \n  '), [], '纯空白')
check_eq(parse_names('张三'), [('', '张三')], '单行单姓名')
check_eq(parse_names('张三\n李四'), [('', '张三'), ('', '李四')], '换行分隔')
check_eq(parse_names('张三,李四'), [('', '张三'), ('', '李四')], '逗号分隔')
check_eq(parse_names('张三,李四,王五'), [('', '张三'), ('', '李四'), ('', '王五')], '逗号多姓名')
check_eq(parse_names('张三、李四'), [('', '张三'), ('', '李四')], '顿号分隔')
check_eq(parse_names('张三；李四'), [('', '张三'), ('', '李四')], '中文分号')
check_eq(parse_names('张三 李四'), [('', '张三'), ('', '李四')], '空格分隔')
check_eq(parse_names('  张三  \n  李四  '), [('', '张三'), ('', '李四')], 'trim 前后空格')
check_eq(parse_names('张三\n\n李四'), [('', '张三'), ('', '李四')], '忽略空行')
check_eq(parse_names('张三，李四'), [('', '张三'), ('', '李四')], '中文逗号')


# ════════════════════════════════════════════
# Test 2: build_schedule 基础
# ════════════════════════════════════════════
section('Test 2: build_schedule')

rows, sec = build_schedule([('', 'A'), ('', 'B')], '08:00', '08:30')
check_eq(len(rows), 2, '2 人 → 2 行')
check_eq(rows[0][0], 1, '第一行序号 1')
check_eq(rows[0][1], '', '第一行学号空')
check_eq(rows[0][2], 'A', '第一行姓名 A')
check_eq(rows[0][3], '08:00', '第一人开始 08:00')
check_eq(rows[0][4], '08:15', '第一人结束 08:15')
check_eq(rows[1][0], 2, '第二行序号 2')
check_eq(rows[1][2], 'B', '第二行姓名 B')
check_eq(rows[1][3], '08:15', '第二人开始 08:15')
check_eq(rows[1][4], '08:30', '第二人结束 08:30')
check_eq(sec, 900.0, '每人 900 秒（15 分钟）')


# ════════════════════════════════════════════
# Test 3: build_schedule 带学号
# ════════════════════════════════════════════
section('Test 3: build_schedule 带学号')

rows_id, sec_id = build_schedule(
    [('04230001', '张三'), ('04230002', '李四')], '10:00', '11:00'
)
check_eq(len(rows_id), 2, '2 行')
check_eq(rows_id[0][1], '04230001', '第一行学号')
check_eq(rows_id[1][1], '04230002', '第二行学号')
check_eq(rows_id[0][2], '张三', '第一行姓名')
check_eq(rows_id[1][2], '李四', '第二行姓名')
check_eq(sec_id, 1800.0, '每人 1800 秒（30 分钟）')


# ════════════════════════════════════════════
# Test 4: build_schedule 边缘
# ════════════════════════════════════════════
section('Test 4: build_schedule 边缘')

rows1, sec1 = build_schedule([('', 'A')], '09:00', '09:01')
check_eq(len(rows1), 1, '1 人 1 分钟 → 1 行')
check_eq(rows1[0][3], '09:00', '开始 09:00')
check_eq(rows1[0][4], '09:01', '结束 09:01')
check_eq(sec1, 60.0, '每人 60 秒')

try:
    build_schedule([('', 'A'), ('', 'B')], '10:00', '10:00')
    check(False, 'end==start 应抛 ValueError')
except ValueError:
    check(True, 'end==start 抛出 ValueError')

try:
    build_schedule([('', 'A'), ('', 'B')], '10:30', '10:00')
    check(False, 'end<start 应抛 ValueError')
except ValueError:
    check(True, 'end<start 抛出 ValueError')

rows0, sec0 = build_schedule([], '08:00', '09:00')
check_eq(len(rows0), 0, '空名单返回空列表')
check_eq(sec0, 0, '空名单 per_student 为 0')


# ════════════════════════════════════════════
# Test 5: build_schedule 多人均匀分配
# ════════════════════════════════════════════
section('Test 5: build_schedule 多人均匀分配')

names = [('', 'A'), ('', 'B'), ('', 'C'), ('', 'D'), ('', 'E')]
rows5, sec5 = build_schedule(names, '08:00', '09:00')
check_eq(len(rows5), 5, '5 人 1 小时 → 5 行')
check_eq(sec5, 720.0, '每人 720 秒（12 分钟）')
for i in range(len(rows5) - 1):
    check(rows5[i][4] == rows5[i+1][3], f'{rows5[i][2]} 结束 == {rows5[i+1][2]} 开始')


# ════════════════════════════════════════════
# Test 6: build_schedule 不整除
# ════════════════════════════════════════════
section('Test 6: build_schedule 不整除')

rows3, sec3 = build_schedule([('', 'A'), ('', 'B'), ('', 'C')], '08:00', '08:01')
check_eq(len(rows3), 3, '3 行')
check_eq(rows3[0][3], '08:00', '第一人开始 08:00')
check(rows3[-1][4] <= '08:01', '最后一人结束 ≤ 08:01')


# ════════════════════════════════════════════
# Test 7: format_duration
# ════════════════════════════════════════════
section('Test 7: format_duration')

check_eq(format_duration(0), '0分钟', '0 秒')
check_eq(format_duration(60), '1分钟', '60 秒')
check_eq(format_duration(120), '2分钟', '120 秒')
check_eq(format_duration(30), '0分30秒', '30 秒')
check_eq(format_duration(90), '1分30秒', '90 秒')
check_eq(format_duration(3661), '61分1秒', '3661 秒')


# ════════════════════════════════════════════
# Test 8: to_excel — 无学号
# ════════════════════════════════════════════
section('Test 8: to_excel（无学号）')

rows_xl, sec_xl = build_schedule(
    [('', '张三'), ('', '李四'), ('', '王五')], '14:00', '15:00'
)
data = to_excel(rows_xl, sec_xl)
check(isinstance(data, bytes), '返回 bytes')
check(len(data) > 0, '非空')

wb = load_workbook(io.BytesIO(data))
check('答辩顺序表' in wb.sheetnames, 'sheet 名称正确')
ws = wb.active
check(ws.cell(1, 1).value == '答辩序号', '表头第一列')
check(ws.cell(1, 2).value == '姓名', '表头第二列为"姓名"（无学号）')
check(ws.cell(2, 1).value == 1, '第一行序号 1')
check(ws.cell(2, 2).value == '张三', '第一行姓名')
wb.close()


# ════════════════════════════════════════════
# Test 9: to_excel — 含学号
# ════════════════════════════════════════════
section('Test 9: to_excel（含学号）')

rows_id2, _ = build_schedule(
    [('04230001', '张三'), ('04230002', '李四')], '14:00', '15:00'
)
data2 = to_excel(rows_id2, 1800.0)
check(isinstance(data2, bytes), '返回 bytes')

wb2 = load_workbook(io.BytesIO(data2))
ws2 = wb2.active
check(ws2.cell(1, 1).value == '答辩序号', '表头第一列')
check(ws2.cell(1, 2).value == '学号', '表头第二列为学号')
check(ws2.cell(1, 3).value == '姓名', '表头第三列为姓名')
check(ws2.cell(2, 2).value == '04230001', '第一行学号')
check(ws2.cell(2, 3).value == '张三', '第一行姓名')
wb2.close()


# ════════════════════════════════════════════
# Test 10: parse_roster
# ════════════════════════════════════════════
section('Test 10: parse_roster')

roster_path = os.path.join(
    os.path.dirname(__file__), '..', '..', '..', 'samples', 'sample_class_roster.xlsx'
)
students = parse_roster(roster_path)
check(len(students) == 10, f'解析出 10 名学生（实际 {len(students)}）')
check_eq(students[0], ('04230001', '张三'), '第 1 人：04230001 张三')
check_eq(students[1], ('04230002', '李四'), '第 2 人：04230002 李四')
check_eq(students[4], ('04230005', '孙七'), '第 5 人：04230005 孙七')
check(all(s[0] for s in students), '所有学生均有学号')
check(all(s[1] for s in students), '所有学生均有姓名')


# ════════════════════════════════════════════
# 汇总
# ════════════════════════════════════════════
section('汇总')
total = PASS + FAIL
print(f'  {PASS}/{total} 通过')
if FAIL:
    print(f'  ❌ {FAIL} 失败')
    sys.exit(1)
else:
    print('  ✅ 全部通过')
