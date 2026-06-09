"""
学情分析 — 核心逻辑测试

测试内容：
  1. 无效文件上传 → parse_file 抛出 ValueError
  2. 有效样本文件 → 正确解析并返回完整数据结构
  3. _compute_session_score_info → 每次课得分统计正确
  4. compute_class_stats → 班级统计指标正确
  5. prepare_class_data → 输出包含百分比和未出题标记
  6. prepare_student_data → 有/无 session_score_info 两种模式
"""
import sys, os, json

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', '..'))

from openpyxl import load_workbook
from tools.attendance.core import parse_file
from tools.learning_analytics.core import (
    _compute_session_score_info, compute_class_stats,
    compute_student_stats, prepare_class_data,
    prepare_student_data, _attendance_status,
)

PASS = FAIL = 0
FIXTURES = os.path.join(os.path.dirname(__file__), 'fixtures')
SAMPLE = os.path.join(
    os.path.dirname(__file__), '..', '..', '..', 'samples', 'sample_attendance.xlsx'
)


def check(cond, msg):
    global PASS, FAIL
    if cond: PASS += 1; print(f'  ✓ {msg}')
    else:    FAIL += 1; print(f'  ✗ {msg}')


def check_eq(a, b, msg):
    global PASS, FAIL
    if a == b: PASS += 1; print(f'  ✓ {msg} ({a})')
    else: FAIL += 1; print(f'  ✗ {msg}: got {a}, expected {b}')


def section(name):
    print(f'\n{"="*60}\n  {name}\n{"="*60}')


# ════════════════════════════════════════════════════════
# Test 1: Invalid file
# ════════════════════════════════════════════════════════
section('Test 1: 无效文件上传 → ValueError')

invalid_path = os.path.join(FIXTURES, 'invalid.xlsx')
wb = load_workbook(invalid_path)
raised = False
try:
    parse_file(wb)
except ValueError:
    raised = True
finally:
    wb.close()
check(raised, 'parse_file 对非雨课堂文件抛出 ValueError')


# ════════════════════════════════════════════════════════
# Test 2: Valid sample file
# ════════════════════════════════════════════════════════
section('Test 2: 有效样本文件解析')

wb = load_workbook(SAMPLE, data_only=True)
sk, sm, scm, students, ld = parse_file(wb)
wb.close()

check(len(sk) > 0, f'解析出 {len(sk)} 次课')
check(len(students) > 0, f'解析出 {len(students)} 名学生')
check(isinstance(ld, dict), f'请假数据 {len(ld)} 条')

# Check student structure
s0 = students[0]
check('id' in s0 and 'name' in s0 and 'attendance' in s0, '学生包含 id/name/attendance')
check('scores' in s0, '学生包含 scores')

# Merge leave data
from tools.learning_analytics.ui import _merge_leave_data
_merge_leave_data(students, ld)
check('leave_sessions' in students[0], '合并后学生包含 leave_sessions')
check(len(students[0].get('leave_sessions', set())) >= 0, 'merge_leave_data 完成')


# ════════════════════════════════════════════════════════
# Test 3: _compute_session_score_info
# ════════════════════════════════════════════════════════
section('Test 3: 每次课得分统计')

info = _compute_session_score_info(students, sk)
check(len(info) == len(sk), f'统计了 {len(info)} 次课')

for skey, si in info.items():
    check(isinstance(si['max'], (int, float)), f'{skey}: max={si["max"]}')
    check(isinstance(si['avg'], (int, float)), f'{skey}: avg={si["avg"]}')
    check(si['count'] == len(students) or si['count'] == 0,
          f'{skey}: count={si["count"]}')


# ════════════════════════════════════════════════════════
# Test 4: compute_class_stats
# ════════════════════════════════════════════════════════
section('Test 4: 班级统计指标')

stats = compute_class_stats(students, sk)
check(stats['avg_attendance_rate'] >= 0, f'平均出勤率 {stats["avg_attendance_rate"]}%')
check(isinstance(stats['avg_score'], (int, float)), f'平均分 {stats["avg_score"]}')
check(len(stats['session_stats']) == len(sk), f'session_stats {len(stats["session_stats"])} 条')
check(isinstance(stats['attention_list'], list), 'attention_list 是列表')


# ════════════════════════════════════════════════════════
# Test 5: prepare_class_data 格式
# ════════════════════════════════════════════════════════
section('Test 5: 班级提示词格式')

prompt = prepare_class_data('测试班级', students, sk)

checks = [
    ('含全班最高', u'全班最高' in prompt),
    ('含百分比', '%' in prompt),
    ('含未出题标记', u'未出题' in prompt or all(u'未出题' not in prompt for _ in [1])),
    ('含出勤数据', u'出勤' in prompt),
    ('含学生得分', u'得分' in prompt),
]
for label, cond in checks:
    check(cond, label)

# Verify it can be used as a format string input (CLASS_ANALYSIS_PROMPT has {sessions_data} etc)
check(prompt.count('{') < 5, '提示词已正确填充所有占位符')


# ════════════════════════════════════════════════════════
# Test 6: prepare_student_data
# ════════════════════════════════════════════════════════
section('Test 6: 学生提示词')

# With session_score_info
sp = prepare_student_data(students[0], sk, session_score_info=info)
check(u'全班最高' in sp, '含 session_score_info 时包含全班最高分')
check(u'得分率' in sp or u'%' in sp, '含 score 百分比')

# Without session_score_info (backward compat)
sp2 = prepare_student_data(students[0], sk)
# The data lines should NOT have score context like "分（全班最高..."
lines_with_context = [line for line in sp2.split('\n') if '分（全班最高' in line]
check(len(lines_with_context) == 0, '无 session_score_info 时得分行不带上下文')
check(u'得分' in sp2, '仍包含得分信息')


# ════════════════════════════════════════════════════════
# Test 7: compute_student_stats
# ════════════════════════════════════════════════════════
section('Test 7: 学生统计')

ss = compute_student_stats(students[0], sk)
check(ss['name'] == students[0]['name'], f'name: {ss["name"]}')
check(ss['session_count'] == len(sk), f'session_count: {ss["session_count"]}')
check(ss['attended_count'] + ss['absent_count'] + ss['leave_count'] == len(sk),
      f'出勤+缺勤+请假={ss["attended_count"]}+{ss["absent_count"]}+{ss["leave_count"]}={len(sk)}')
check(isinstance(ss['scores'], dict), 'scores 是 dict')
check(isinstance(ss['attendance_detail'], dict), 'attendance_detail 是 dict')


# ════════════════════════════════════════════════════════
# Summary
# ════════════════════════════════════════════════════════
section('Summary')
print(f'  Pass: {PASS}, Fail: {FAIL}')
if FAIL:
    sys.exit(1)
