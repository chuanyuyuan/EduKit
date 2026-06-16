"""
学情分析 — Streamlit UI 组件
"""
import os
import base64

import streamlit as st
import pandas as pd
from openpyxl import load_workbook

from tools.attendance.core import parse_file
from .core import (
    analyze_student, compute_class_stats, compute_student_stats,
    call_deepseek, prepare_class_data, _attendance_status,
    _compute_session_score_info,
)
from .prompts import CLASS_ANALYSIS_SYSTEM

DEMO_FILE = "samples/sample_attendance.xlsx"


def _demo_path():
    return os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        DEMO_FILE,
    )


def _sample_link():
    """Return base64 data URI for the demo file download link."""
    path = _demo_path()
    if not os.path.exists(path):
        return ""
    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode()
    return f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}"


def _get_api_key():
    """获取 DeepSeek API Key：st.secrets（部署）→ 环境变量 / config（本地）。"""
    try:
        key = st.secrets.get("DEEPSEEK_API_KEY", "")
        if key:
            return key
    except Exception:
        pass
    from .config import DEEPSEEK_API_KEY
    return DEEPSEEK_API_KEY


def _need_api_warning():
    if not _get_api_key():
        st.warning("未配置 DeepSeek API Key，AI 分析功能不可用。部署后在 Streamlit Cloud Dashboard → Secrets 中配置 DEEPSEEK_API_KEY，或本地设置环境变量。统计指标仍正常显示。", icon="⚠️")


def render_page():
    st.header(":material/insights: 雨课堂学情分析")

    link = _sample_link()
    st.markdown(f"""
上传雨课堂导出的 Excel 文件，自动分析班级出勤与得分趋势，生成 AI 学情评语和学生画像。

**适用平台：** [长江雨课堂](https://changjiang.yuketang.cn/web/?index)

**支持的文件类型：**
- 批量导出的汇总数据表（含汇总页 + 课堂情况子表）
- 单次课导出的课堂数据表（自动识别）

**原始文件获取方式：**
1. 登录长江雨课堂官网，进入你授课的班级
2. 点击 **批量导出数据**
3. 筛选需要统计的课堂记录
4. 下载导出的 Excel 文件（即 `.xlsx` 格式的汇总数据表）。<a href="{link}" download="{DEMO_FILE}">下载示例表格</a>

**上传后：**
- 自动解析并分 tab 展示班级总览和学生画像
- 支持下载分析报告（Excel）
""", unsafe_allow_html=True)

    _need_api_warning()

    # ── 文件上传 / 示例数据 ──
    uploaded = st.file_uploader("选择雨课堂导出的 Excel 文件（仅支持 .xlsx 格式）")

    result_key = "learning_analytics_result"

    if "la_show_demo" not in st.session_state:
        st.session_state.la_show_demo = False

    if not uploaded and not st.session_state.la_show_demo:
        if st.button(":material/lightbulb: 加载示例数据看看效果", use_container_width=True):
            st.session_state.la_show_demo = True
            st.rerun()

    # ── 输入源变化时清除缓存 ──
    _last_file = st.session_state.get("la_prev_file")
    if uploaded:
        if uploaded.name != _last_file:
            st.session_state.pop(result_key, None)
            st.session_state.la_show_demo = False
    elif _last_file is not None:
        # 叉掉了上传文件 或 切换到演示模式
        st.session_state.pop(result_key, None)
        st.session_state.la_prev_file = None
        if not st.session_state.la_show_demo:
            # 叉掉文件后回到初始状态
            st.rerun()

    # ── 解析 ──
    if st.session_state.get(result_key) is None:
        session_keys = sign_map = score_map = students = None

        if uploaded:
            if not uploaded.name.endswith('.xlsx'):
                st.error(f'不支持的文件格式："{uploaded.name}"，请上传 .xlsx 文件。')
            else:
                wb_path = None
                try:
                    wb_path = _save_upload(uploaded)
                    wb = load_workbook(wb_path, data_only=True)
                    session_keys, sign_map, score_map, students, leave_data = parse_file(wb)
                    wb.close()
                    _merge_leave_data(students, leave_data)
                except ValueError as e:
                    st.error(f"不支持该文件格式。请上传雨课堂批量导出的汇总数据表或单次课数据表。\n\n错误详情：{e}")
                finally:
                    if wb_path:
                        try:
                            os.remove(wb_path)
                        except OSError:
                            pass
        elif st.session_state.la_show_demo:
            demo = _demo_path()
            if os.path.exists(demo):
                with st.status("正在解析...", expanded=True) as status:
                    wb = load_workbook(demo, data_only=True)
                    session_keys, sign_map, score_map, students, leave_data = parse_file(wb)
                    wb.close()
                    _merge_leave_data(students, leave_data)
                    status.update(label=f"解析完成：{len(students)} 名学生，{len(session_keys)} 次课", state="complete")

        if students:
            with st.status("正在分析...", expanded=True) as status:
                # 统计指标（快速）
                class_result = {"stats": compute_class_stats(students, session_keys)}
                status.write("✓ 统计指标计算完成")

                # 班级 AI 分析
                status.write("→ 正在调用 AI 分析班级数据...")
                _api_key = _get_api_key()
                ai_class = call_deepseek(CLASS_ANALYSIS_SYSTEM, prepare_class_data(
                    "示例班级" if st.session_state.la_show_demo else uploaded.name,
                    students, session_keys,
                ), api_key=_api_key)
                if "error" not in ai_class:
                    class_result["ai"] = ai_class
                status.write("✓ 班级 AI 分析完成")

                # 预生成所有学生 AI 评语（带进度条 + 预估时间）
                import time
                session_score_info = _compute_session_score_info(students, session_keys)
                student_ai_cache = {}
                n = len(students)
                progress = st.progress(0, text="正在生成学生 AI 评语...")
                t_start = time.time()
                for i, s in enumerate(students):
                    sid = s.get("id", "") or s.get("name", "")
                    sr = analyze_student(s, session_keys, session_score_info, api_key=_api_key)
                    ai = sr.get("ai")
                    if ai and "error" not in ai:
                        student_ai_cache[sid] = ai
                    # 计算预估剩余时间
                    elapsed = time.time() - t_start
                    avg = elapsed / (i + 1)
                    remaining = avg * (n - i - 1)
                    eta = f"{remaining:.0f}秒" if remaining < 60 else f"{remaining / 60:.0f}分{remaining % 60:.0f}秒"
                    progress.progress(
                        (i + 1) / n,
                        text=f"正在生成学生 AI 评语 ({i + 1}/{n}) · 预计剩余 {eta}",
                    )
                progress.empty()
                status.write(f"✓ 全部 {n} 名学生 AI 评语生成完成")
                status.update(label=f"分析完成：{n} 名学生", state="complete", expanded=False)
            st.session_state[result_key] = {
                "class_result": class_result,
                "students": students,
                "session_keys": session_keys,
                "sign_map": sign_map,
                "score_map": score_map,
                "student_ai_cache": student_ai_cache,
            }
            st.session_state.la_prev_file = uploaded.name if uploaded else None
            st.rerun()

    result = st.session_state.get(result_key)
    if not result:
        if not uploaded and not st.session_state.la_show_demo:
            st.info("请上传雨课堂 Excel 文件开始分析。")
        return

    # ════════════════════════════════════════
    #  展示结果
    # ════════════════════════════════════════
    class_result = result["class_result"]
    students = result["students"]
    session_keys = result["session_keys"]

    st.markdown("---")

    tab1, tab2 = st.tabs([":material/group: 班级总览", ":material/person: 学生画像"])

    with tab1:
        _render_class_overview(class_result, session_keys)

    with tab2:
        _render_student_profile(students, session_keys, result.get("student_ai_cache", {}))

    # 下载报告
    col_left, _ = st.columns([1, 4])
    with col_left:
        st.download_button(
            ":material/download: 下载分析报告",
            data=_generate_report(class_result, students, session_keys, result.get("student_ai_cache", {})),
            file_name="学情分析报告.xlsx",
            use_container_width=True,
        )

    # 示例数据提示
    if st.session_state.la_show_demo:
        st.info("当前展示的是示例数据，你可以上传自己的 Excel 文件替换。")


def _save_upload(uploaded):
    """将上传文件保存到临时路径。"""
    import tempfile
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    with open(fd, "wb") as f:
        f.write(uploaded.getvalue())
    return path


def _merge_leave_data(students, leave_data):
    """将请假数据合并到学生字典中。"""
    for s in students:
        sid = s.get("id", "")
        s["leave_sessions"] = set()
        for (l_sk, l_sid), l_type in leave_data.items():
            if l_sid == sid:
                s["leave_sessions"].add(l_sk)


def _generate_report(class_result, students, session_keys, student_ai_cache=None):
    """生成带样式的分析报告 Excel 文件（内存中），返回 bytes。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    light_blue_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(1, col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

    def style_data_rows(ws, start_row, end_row, ncols):
        for r in range(start_row, end_row + 1):
            for c in range(1, ncols + 1):
                cell = ws.cell(r, c)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if r % 2 == 0:
                    cell.fill = light_blue_fill

    # ── Sheet 1: 班级统计 ──
    ws1 = wb.active
    ws1.title = "班级统计"
    ws1.append(["课次", "出勤率(%)", "平均分"])
    for row in class_result["stats"]["session_stats"]:
        ws1.append([row["session"], row["attendance_rate"], row["avg_score"]])
    ws1.append([])
    summary_start = ws1.max_row + 1
    ws1.append(["平均出勤率", class_result["stats"]["avg_attendance_rate"], ""])
    ws1.append(["平均得分", "", class_result["stats"]["avg_score"]])
    style_header(ws1, 3)
    style_data_rows(ws1, 2, len(class_result["stats"]["session_stats"]) + 1, 3)
    # 汇总行加粗
    for r in range(summary_start, ws1.max_row + 1):
        for c in range(1, 4):
            cell = ws1.cell(r, c)
            cell.font = Font(bold=True, size=11)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 16

    # ── Sheet 2: 需关注学生 ──
    ws2 = wb.create_sheet("需关注学生")
    _ai = class_result.get("ai")
    if _ai and "error" not in _ai and _ai.get("attention_students"):
        ws2.append(["学生", "原因"])
        for item in _ai["attention_students"]:
            ws2.append(["", item])
    else:
        ws2.append(["学生", "原因"])
        for name, reason in class_result["stats"]["attention_list"]:
            ws2.append([name, reason])
    if ws2.max_row > 1:
        style_header(ws2, 2)
        style_data_rows(ws2, 2, ws2.max_row, 2)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 40

    # ── Sheet 3: 学生详情 ──
    ws3 = wb.create_sheet("学生详情")
    headers = ["学号", "姓名", "出勤率(%)", "出勤次数", "缺勤次数", "请假次数"]
    headers += [f"{sk}(得分)" for sk in session_keys]
    headers += [f"{sk}(出勤)" for sk in session_keys]
    ws3.append(headers)
    ncols = len(headers)

    for s in students:
        sid = s.get("id", "")
        name = s.get("name", "")
        att_count = sum(
            1 for sk in session_keys
            if _attendance_status(
                s.get("attendance", {}).get(sk, ""),
                sk in s.get("leave_sessions", set()),
            ) == "上课"
        )
        absent_count = sum(
            1 for sk in session_keys
            if _attendance_status(
                s.get("attendance", {}).get(sk, ""),
                sk in s.get("leave_sessions", set()),
            ) == "缺勤"
        )
        leave_count = sum(
            1 for sk in session_keys
            if _attendance_status(
                s.get("attendance", {}).get(sk, ""),
                sk in s.get("leave_sessions", set()),
            ) == "请假"
        )
        rate = round(att_count / len(session_keys) * 100, 1) if session_keys else 0
        row = [sid, name, rate, att_count, absent_count, leave_count]
        for sk in session_keys:
            row.append(s.get("scores", {}).get(sk, ""))
        for sk in session_keys:
            row.append(
                _attendance_status(
                    s.get("attendance", {}).get(sk, ""),
                    sk in s.get("leave_sessions", set()),
                )
            )
        ws3.append(row)

    style_header(ws3, ncols)
    style_data_rows(ws3, 2, ws3.max_row, ncols)
    # 出勤率低的行高亮
    for r in range(2, ws3.max_row + 1):
        rate_cell = ws3.cell(r, 3)
        if rate_cell.value is not None and isinstance(rate_cell.value, (int, float)) and rate_cell.value < 70:
            for c in range(1, ncols + 1):
                ws3.cell(r, c).fill = yellow_fill

    from openpyxl.utils import get_column_letter
    for i in range(1, ncols + 1):
        ws3.column_dimensions[get_column_letter(i)].width = 18 if i <= 2 else 14

    # ── Sheet 4: AI 分析结果 ──
    ai = class_result.get("ai")
    if ai and "error" not in ai:
        ws4 = wb.create_sheet("AI 分析")
        ws4.append(["指标", "内容"])
        ws4.cell(1, 1).font = Font(bold=True, size=11)
        ws4.cell(1, 2).font = Font(bold=True, size=11)
        rows = [
            ("班级评语", ai.get("summary", "")),
            ("AI 出勤率", ai.get("attendance_rate", "")),
            ("AI 平均分", ai.get("avg_score", "")),
            ("趋势", ai.get("trend", "")),
        ]
        for name, val in rows:
            ws4.append([name, val])
        anomalies = ai.get("anomalies", [])
        if anomalies:
            ws4.append([])
            ws4.append(["异常检测", ""])
            ws4.cell(ws4.max_row, 1).font = Font(bold=True)
            for a in anomalies:
                ws4.append(["", a])
        attention = ai.get("attention_students", [])
        if attention:
            ws4.append([])
            ws4.append(["AI 建议关注", ""])
            ws4.cell(ws4.max_row, 1).font = Font(bold=True)
            for a in attention:
                ws4.append(["", a])
        ws4.column_dimensions["A"].width = 18
        ws4.column_dimensions["B"].width = 55

    # ── Sheet 5: 学生 AI 评语 ──
    if student_ai_cache:
        ws5 = wb.create_sheet("学生 AI 评语")
        ws5.append(["学号/姓名", "评语", "优势", "薄弱点", "建议"])
        for cell in ws5[1]:
            cell.font = Font(bold=True, size=11)
        for sid_or_name in sorted(student_ai_cache.keys()):
            s = student_ai_cache[sid_or_name]
            ws5.append([
                sid_or_name,
                s.get("comment", ""),
                s.get("strength", ""),
                s.get("weakness", ""),
                s.get("suggestion", ""),
            ])
        ws5.column_dimensions["A"].width = 16
        ws5.column_dimensions["B"].width = 40
        ws5.column_dimensions["C"].width = 30
        ws5.column_dimensions["D"].width = 30
        ws5.column_dimensions["E"].width = 35

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════
# 班级总览
# ════════════════════════════════════════════════════════════

def _render_class_overview(class_result, session_keys):
    stats = class_result["stats"]
    ai = class_result.get("ai")
    has_ai = ai and "error" not in ai

    # 指标卡片
    m1, m2, m3, m4 = st.columns(4)
    m1.metric(":material/event: 总课次", f"{len(session_keys)} 次")
    m2.metric(":material/check_circle: 平均出勤率", f"{stats['avg_attendance_rate']}%")
    m3.metric(":material/star: 平均得分", f"{stats['avg_score']}")
    attention_count = len(ai["attention_students"]) if (has_ai and ai.get("attention_students")) else len(stats['attention_list'])
    m4.metric(":material/priority_high: 需关注人数", f"{attention_count}")

    # 趋势图
    if stats["session_stats"]:
        st.subheader("每次课趋势")
        df = pd.DataFrame(stats["session_stats"])
        try:
            import altair as alt
            chart1 = alt.Chart(df).mark_line(point=True).encode(
                x=alt.X("session:N", title="课次", sort=None),
                y=alt.Y("attendance_rate:Q", title="出勤率(%)", scale=alt.Scale(domain=[0, 100])),
            ).properties(height=200)
            st.altair_chart(chart1, use_container_width=True)

            chart2 = alt.Chart(df).mark_line(point=True, color="#E74C3C").encode(
                x=alt.X("session:N", title="课次", sort=None),
                y=alt.Y("avg_score:Q", title="平均分"),
            ).properties(height=200)
            st.altair_chart(chart2, use_container_width=True)
        except ImportError:
            st.dataframe(df, hide_index=True)

    # 需关注学生（优先用 AI，AI 不可用时回落规则计算）
    attention_items = None
    if has_ai and ai.get("attention_students"):
        attention_items = ai["attention_students"]
    elif stats["attention_list"]:
        attention_items = [f"{name}（{reason}）" for name, reason in stats["attention_list"]]

    if attention_items:
        st.subheader(":material/priority_high: 需关注学生")
        for item in attention_items:
            st.write(f"- {item}")

    # AI 学情分析
    if has_ai:
        st.subheader(":material/psychology: AI 学情分析")

        if ai.get("summary"):
            st.info(ai["summary"])

        cols = st.columns(2)
        cols[0].metric("AI 出勤率", ai.get("attendance_rate", ""))
        cols[1].metric("AI 平均分", ai.get("avg_score", ""))

        trend = ai.get("trend", "")
        if trend:
            st.markdown(f"**趋势**：{trend}")

        if ai.get("anomalies"):
            with st.expander("异常检测", expanded=True):
                for a in ai["anomalies"]:
                    st.write(f"- {a}")


# ════════════════════════════════════════════════════════════
# 学生画像
# ════════════════════════════════════════════════════════════

def _render_student_profile(students, session_keys, student_ai_cache=None):
    if not students:
        return

    names = [s.get("name", f"学生{i}") for i, s in enumerate(students)]
    selected = st.selectbox("选择学生", names, label_visibility="collapsed")

    student = students[names.index(selected)]
    session_score_info = _compute_session_score_info(students, session_keys)
    stats = compute_student_stats(student, session_keys)

    # 优先用缓存中的 AI 评语，缓存未命中时实时调用
    sid = student.get("id", "") or student.get("name", "")
    ai = None
    if student_ai_cache and sid in student_ai_cache:
        ai = student_ai_cache[sid]
    else:
        ai = analyze_student(student, session_keys, session_score_info, api_key=_get_api_key()).get("ai")

    # 基本信息
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("学号", stats["student_id"])
    col2.metric("出勤率", f"{stats['attendance_rate']}%")
    col3.metric("平均得分", _avg_score(stats))
    col4.metric("缺勤", f"{stats['absent_count']} 次")

    # 雷达图（出勤率, 得分率, 出勤稳定性）
    _draw_radar(stats, session_score_info)

    # 得分趋势（得分率）
    if stats["scores"] and session_score_info:
        st.subheader("得分率趋势")
        df_rate = []
        for sk, sc in stats["scores"].items():
            si = session_score_info.get(sk)
            rate = round(sc / si["max"] * 100, 1) if si and si.get("max", 0) > 0 else None
            df_rate.append({"课次": sk, "得分率(%)": rate})
        df_rate = pd.DataFrame(df_rate)
        try:
            import altair as alt
            chart = alt.Chart(df_rate).mark_line(point=True, color="#E74C3C").encode(
                x=alt.X("课次:N", sort=None),
                y=alt.Y("得分率(%):Q", scale=alt.Scale(domain=[0, 100])),
            ).properties(height=200)
            st.altair_chart(chart, use_container_width=True)
        except ImportError:
            st.dataframe(df_rate, hide_index=True)
    elif stats["scores"]:
        st.subheader("得分趋势")
        df_score = pd.DataFrame([
            {"课次": sk, "得分": sc}
            for sk, sc in stats["scores"].items()
        ])
        st.dataframe(df_score, hide_index=True)

    # 出勤明细
    st.subheader("出勤明细")
    detail_df = pd.DataFrame([
        {"课次": sk, "状态": st}
        for sk, st in stats["attendance_detail"].items()
    ])
    st.dataframe(detail_df, use_container_width=True, hide_index=True)

    # AI 评语
    if ai and "error" not in ai:
        st.subheader(":material/psychology: AI 学情评语")
        if ai.get("comment"):
            st.info(ai["comment"])
        cols = st.columns(3)
        if ai.get("strength"):
            cols[0].success(f"优势\n{ai['strength']}")
        if ai.get("weakness"):
            cols[1].warning(f"薄弱点\n{ai['weakness']}")
        if ai.get("suggestion"):
            cols[2].info(f"建议\n{ai['suggestion']}")


def _avg_score(stats):
    scores = list(stats["scores"].values())
    if scores:
        return f"{sum(scores) / len(scores):.1f}"
    return "-"


def _draw_radar(stats, session_score_info=None):
    """用 matplotlib 画简易雷达图。

    三个维度（均归一化到 0-100）：
    - 出勤率
    - 得分水平（平均得分率，需 session_score_info 计算）
    - 出勤稳定性（基于缺勤占比）
    """
    scores = stats["scores"]

    # 得分水平：各次课得分率的均值
    score_level = 0
    if session_score_info and scores:
        rates = []
        for sk, score in scores.items():
            si = session_score_info.get(sk)
            if si and si.get("max", 0) > 0:
                rates.append(score / si["max"] * 100)
        score_level = sum(rates) / len(rates) if rates else 0
    elif scores:
        sv = list(scores.values())
        score_level = sum(sv) / len(sv) if sv else 0

    # 出勤稳定性：基于缺勤占比（请假视为部分稳定）
    total = stats["session_count"]
    absent = stats["absent_count"]
    leave = stats["leave_count"]
    if total:
        stability = (1 - (absent + 0.5 * leave) / total) * 100
    else:
        stability = 100

    labels = ["出勤率", "得分水平", "出勤稳定性"]
    values = [stats["attendance_rate"], score_level, max(0, stability)]

    try:
        from matplotlib import font_manager
        import matplotlib.pyplot as plt
        import numpy as np

        _cn_candidates = [f.name for f in font_manager.fontManager.ttflist
                          if any(k in f.name.lower() for k in ("yahei", "simhei", "noto", "wenquanyi",
                                                               "dengxian", "stxihei", "stsong",
                                                               "arial unicode"))]
        rcParams = plt.rcParams
        rcParams['font.sans-serif'] = _cn_candidates + ['DejaVu Sans', 'sans-serif']
        rcParams['axes.unicode_minus'] = False

        angles = np.linspace(0, 2 * np.pi, len(labels), endpoint=False).tolist()
        values += values[:1]
        angles += angles[:1]

        fig, ax = plt.subplots(figsize=(3, 3), subplot_kw={"polar": True})
        ax.plot(angles, values, "o-", linewidth=2, color="#3498DB")
        ax.fill(angles, values, alpha=0.15, color="#3498DB")
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(labels, fontsize=9)
        ax.set_ylim(0, 100)
        ax.set_yticks([25, 50, 75])
        ax.set_yticklabels(["", "", ""])
        st.pyplot(fig, use_container_width=False)
        plt.close(fig)
    except ImportError:
        pass
