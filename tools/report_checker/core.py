"""
实验报告查重 — 核心逻辑

流水线：解压 ZIP → 整理 Word 文档 → 图片像素 MD5 指纹提取 →
交叉比对 → Excel 报表生成

.docx → python-docx 读取嵌入图片
.doc → olefile 从 OLE2 Data 流中扫描 PNG，Pillow 解码到像素后计算 MD5
（消除 WPS Writer 元数据干扰，纯 Python，无需 Word/pywin32）
"""
import os
import shutil
import hashlib
import zipfile
from io import BytesIO

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint


def unzip_process(zip_path, raw_dir):
    """解压 ZIP 到 raw_dir。尝试 cp437→utf-8 修复旧版 ZIP 中文乱码。"""
    os.makedirs(raw_dir, exist_ok=True)
    with zipfile.ZipFile(zip_path, 'r') as z:
        for m in z.infolist():
            raw = m.filename
            try:
                name = raw.encode('cp437').decode('utf-8')
            except UnicodeEncodeError:
                # 含有 cp437 无法编码的字符 → 已是正确 Unicode
                name = raw
            except Exception:
                # cp437→utf-8 修复失败 → 保留原文件名
                name = raw
            m.filename = name
            z.extract(m, raw_dir)


def organize_process(raw_dir, target_dir):
    """遍历学生子目录，提取第一份非「答题记录」的 Word 报告，重命名。

    对于无实验报告的学生（如仅含答题记录的目录），创建空白 .docx 占位，
    确保该学生在后续分析中仍会被统计为「无图片」。
    """
    os.makedirs(target_dir, exist_ok=True)
    folders = [f for f in os.listdir(raw_dir) if os.path.isdir(os.path.join(raw_dir, f))]
    for folder in folders:
        f_path = os.path.join(raw_dir, folder)
        parts = folder.split('_')
        new_name = f"{parts[0]}_{parts[1]}" if len(parts) >= 2 else folder
        found = False
        for file in os.listdir(f_path):
            if "答题记录" not in file and not file.startswith('~$'):
                ext = os.path.splitext(file)[1].lower()
                if ext in ('.doc', '.docx'):
                    shutil.copy(
                        os.path.join(f_path, file),
                        os.path.join(target_dir, f"{new_name}{ext}"),
                    )
                    found = True
                    break
        if not found:
            Document().save(os.path.join(target_dir, f"{new_name}.docx"))


def extract_features(file_path):
    """提取 doc/docx 中嵌入图片的 (MD5, 字节大小) 指纹列表。

    使用 Pillow 解码到像素数据后计算 MD5，消除 WPS Writer/Word
    在 PNG 元数据中嵌入的时间戳等差异，确保相同截图产生相同指纹。
    """
    ext = os.path.splitext(file_path)[1].lower()
    blobs = []

    if ext == '.docx':
        try:
            doc = Document(file_path)
            blobs = [(r.target_part.blob, len(r.target_part.blob))
                     for r in doc.part.rels.values() if "image" in r.target_ref]
        except Exception:
            return []
    elif ext == '.doc':
        try:
            import olefile
            ole = olefile.OleFileIO(file_path)
            if not ole.exists('Data'):
                ole.close()
                return []
            data = ole.openstream('Data').read()
            ole.close()
            start = 0
            png_magic = b'\x89PNG\r\n\x1a\n'
            while True:
                pos = data.find(png_magic, start)
                if pos == -1:
                    break
                iend = data.find(b'IEND', pos)
                if iend == -1:
                    start = pos + 1
                    continue
                png_data = data[pos:iend + 12]
                blobs.append((png_data, len(png_data)))
                start = pos + 1
        except ImportError:
            return []
        except Exception:
            return []
    else:
        return []

    # 解码到像素后计算 MD5，消除元数据干扰
    from PIL import Image
    import io
    result = []
    for blob, size in blobs:
        try:
            img = Image.open(io.BytesIO(blob))
            pixel_md5 = hashlib.md5(img.tobytes()).hexdigest()
            result.append((pixel_md5, size))
        except Exception:
            continue
    return result


def analyze(target_dir, progress_callback=None):
    """对目标目录下所有 docx 进行指纹交叉比对。

    Args:
        target_dir: 整理后的实验报告目录
        progress_callback: (current, total, message) 进度回调

    Returns:
        dict 含：results（列表）, total, plag_count, none_count
    """
    files = [f for f in os.listdir(target_dir)
             if not f.startswith('~$') and f.lower().endswith(('.doc', '.docx'))]
    n = len(files)
    data = {}
    for i, f in enumerate(files):
        if progress_callback:
            progress_callback(i, 2 * n, f"提取指纹 ({i + 1}/{n}): {f}")
        data[f] = extract_features(os.path.join(target_dir, f))

    results = []
    plag_count = none_count = 0
    f_list = list(data.keys())
    edge_dict = {}  # (name1, name2) -> shared_count

    def _name(f):
        return f.rsplit('.', 1)[0]

    for i in range(len(f_list)):
        if progress_callback:
            progress_callback(n + i, 2 * n,
                              f"交叉比对: 第 {i + 1}/{len(f_list)} 份 — {_name(f_list[i])}")
        f1 = f_list[i]
        h1 = {md5 for md5, _ in data[f1]}
        if not h1:
            results.append([_name(f1), "无图片",
                            "文档内未检测到截图（异常）"])
            none_count += 1
        else:
            matches = []
            for j in range(len(f_list)):
                if i == j:
                    continue
                c = len(h1 & {md5 for md5, _ in data[f_list[j]]})
                if c >= 1:
                    matches.append((c, _name(f_list[j])))
                    key = tuple(sorted([_name(f_list[i]), _name(f_list[j])]))
                    edge_dict[key] = max(edge_dict.get(key, 0), c)
            if matches:
                matches.sort(reverse=True)
                parts = [f"与 [{m[1]}] 相同图片 {m[0]} 张" for m in matches]
                detail = f"报告总图片 {len(h1)} 张；" + "；".join(parts)
                has_plag = any(c / len(h1) >= 0.4 for c, _ in matches)
                results.append([_name(f1), "疑似抄袭" if has_plag else "正常", detail])
                if has_plag:
                    plag_count += 1
            else:
                results.append([_name(f1), "正常",
                                f"报告总图片 {len(h1)} 张"])

    results.sort()
    pairs = [(n1, n2, c) for (n1, n2), c in edge_dict.items()]
    img_counts = {_name(f): len(data[f]) for f in files}
    return {"results": results, "total": len(files),
            "plag_count": plag_count, "none_count": none_count,
            "pairs": pairs, "img_counts": img_counts}


def generate_excel(results, total, plag_count, none_count, pairs=None, img_counts=None):
    """生成查重报告 Excel，返回 BytesIO。"""
    normal_count = total - plag_count - none_count

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "统计概览"
    for r in [["分类项", "人数"],
              ["疑似抄袭", plag_count],
              ["无图异常", none_count],
              ["正常报告", normal_count]]:
        ws1.append(r)
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    chart = PieChart()
    chart.add_data(Reference(ws1, min_col=2, min_row=1, max_row=4),
                   titles_from_data=True)
    chart.set_categories(Reference(ws1, min_col=1, min_row=2, max_row=4))
    chart.title = "实验报告诚信分布"
    for i, c in enumerate(["FFC7CE", "FFEB9C", "C6EFCE"]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = c
        chart.series[0].dPt.append(pt)
    ws1.add_chart(chart, "D2")

    ws2 = wb.create_sheet("详细名单")
    ws2.append(["学号姓名", "判定结论", "详细依据分析"])

    # 无图片排在顶端，其余按学号字母序
    none_rows = [r for r in results if r[1] == "无图片"]
    other_rows = [r for r in results if r[1] != "无图片"]
    sorted_results = none_rows + other_rows
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                           fill_type="solid")
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                              fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                              fill_type="solid")
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(sorted_results, 2):
        ws2.append(row)
        if row[1] == "疑似抄袭":
            for cell in ws2[r_idx]:
                cell.fill = red_fill
        elif row[1] == "无图片":
            for cell in ws2[r_idx]:
                cell.fill = orange_fill
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['C'].width = 65

    # 抄袭关系数据（文字版，网络图无法在 Excel 展示）
    ws3 = wb.create_sheet("抄袭关系")
    ws3.append(["学生A", "学生B", "相同图片数", "抄袭比例(%)"])
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    if pairs:
        for n1, n2, count in pairs:
            total1 = img_counts.get(n1, 1) if img_counts else 1
            total2 = img_counts.get(n2, 1) if img_counts else 1
            ratio = max(count / total1, count / total2) * 100
            ws3.append([n1, n2, count, round(ratio, 1)])
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def run_pipeline(zip_path, workspace, log_func=print, progress_callback=None):
    """完整查重流水线：解压 → 整理 → 分析 → 报表。

    Args:
        zip_path: ZIP 文件路径
        workspace: 临时工作目录
        log_func: 日志回调函数
        progress_callback: (current, total, message) 进度回调，用于 UI 进度条

    Returns:
        dict：results, total, plag_count, none_count, excel_buf
    """
    raw_dir = os.path.join(workspace, "解压缩文件")
    target_dir = os.path.join(workspace, "整理后实验报告")

    log_func("解压文件中...")
    unzip_process(zip_path, raw_dir)

    log_func("提取实验报告...")
    organize_process(raw_dir, target_dir)

    log_func("提取图片指纹并交叉比对...")
    analysis = analyze(target_dir, progress_callback=progress_callback)
    if analysis["total"] == 0:
        log_func("  未找到可分析的文档")

    log_func("生成报表...")
    excel_buf = generate_excel(
        analysis["results"], analysis["total"],
        analysis["plag_count"], analysis["none_count"],
        pairs=analysis.get("pairs"), img_counts=analysis.get("img_counts"),
    )
    return {**analysis, "excel_buf": excel_buf}
