"""
查重分析 — 基于合成数据的端到端测试

使用 tests/sample_report_checker.zip（10 名学生合成样本）验证：
  - 解压中文目录名
  - 文件整理（跳过答题记录）
  - 指纹提取与交叉比对
  - 组 A（4 人共享 6 张图）= 疑似抄袭；组 B（2 人对抄 5 张）= 疑似抄袭
  - 无图片学生（陈十十）= 空白占位，统计为「无图片」
"""
import sys, os, tempfile
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', '..'))

from tools.report_checker.core import (
    unzip_process, organize_process, analyze,
    extract_features, run_pipeline,
)

PASS = FAIL = 0

ZIP_PATH = os.path.join(os.path.dirname(__file__), '..', '..', '..',
                        'tests', 'sample_report_checker.zip')


def check(cond, msg):
    global PASS, FAIL
    if cond: PASS += 1; print(f'  ✓ {msg}')
    else: FAIL += 1; print(f'  ✗ {msg}')


def check_eq(a, b, msg):
    global PASS, FAIL
    if a == b: PASS += 1; print(f'  ✓ {msg} ({a})')
    else: FAIL += 1; print(f'  ✗ {msg}: got {a}, expected {b}')


def section(name):
    print(f'\n{"="*60}\n  {name}\n{"="*60}')


# ════════════════════════════════════════════
section('Test 1: unzip_process — 中文目录名与文件提取')
# ════════════════════════════════════════════

with tempfile.TemporaryDirectory() as td:
    raw = os.path.join(td, "raw")
    unzip_process(ZIP_PATH, raw)
    folders = sorted(os.listdir(raw))
    check_eq(len(folders), 10, f'解压后应有 10 个学生目录（实际 {len(folders)}）')

    # 目录名应为完整中文学号_姓名格式
    chinese_folders = [f for f in folders if any('一' <= c <= '鿿' for c in f)]
    check(len(chinese_folders) == 10, f'所有 10 个目录含中文字符')

    # 所有目录应含答题记录；除陈十十外都应含实验报告
    for f in folders:
        path = os.path.join(raw, f)
        if not os.path.isdir(path):
            continue
        files = os.listdir(path)
        word_files = [x for x in files if not x.startswith('~$')
                      and x.lower().endswith(('.doc', '.docx'))]
        answer_records = [x for x in word_files if '答题记录' in x]
        reports = [x for x in word_files if '答题记录' not in x]
        check(len(answer_records) >= 1, f'{f} 包含答题记录文件')
        if '2024010_陈十十' in f:
            check(len(reports) == 0, f'{f} 无实验报告（合成样本设计）')
        else:
            check(len(reports) >= 1, f'{f} 包含实验报告')


# ════════════════════════════════════════════
section('Test 2: organize_process — 非答题记录报告提取')
# ════════════════════════════════════════════

with tempfile.TemporaryDirectory() as td:
    raw = os.path.join(td, "raw")
    target = os.path.join(td, "organized")
    unzip_process(ZIP_PATH, raw)
    organize_process(raw, target)

    files = sorted(os.listdir(target))
    check_eq(len(files), 10, '整理后共 10 个文件（含陈十十空白占位）')

    # 不应有包含「答题记录」的文件
    answer_in_target = [f for f in files if '答题记录' in f]
    check_eq(len(answer_in_target), 0, '整理结果不含答题记录文件')

    # 文件名格式应为 学号_姓名.ext
    valid_names = [os.path.splitext(f)[0] for f in files]
    double_part = [n for n in valid_names if len(n.split('_')) >= 2]
    check(len(double_part) == len(valid_names),
          f'所有文件名包含学号_姓名格式 ({len(double_part)}/{len(valid_names)})')

    docx_count = len([f for f in files if f.endswith('.docx')])
    doc_count = len([f for f in files if f.endswith('.doc')])
    check_eq(doc_count, 0, '合成样本全部为 .docx，无 .doc 文件')
    check_eq(docx_count, 10, '全部 10 个文件均为 .docx（含陈十十占位）')


# ════════════════════════════════════════════
section('Test 3: analyze — 可读的 .docx 比对')
# ════════════════════════════════════════════

with tempfile.TemporaryDirectory() as td:
    raw = os.path.join(td, "raw")
    target = os.path.join(td, "organized")
    unzip_process(ZIP_PATH, raw)
    organize_process(raw, target)
    result = analyze(target)

    check_eq(result['total'], 10, '合成样本共 10 份文档（含陈十十空白占位）')
    check_eq(result['plag_count'], 6, '抄袭 6 人（组 A 四人 + 组 B 二人）')
    check_eq(result['none_count'], 1, '无图片 1 人（陈十十空白占位）')

    for row in result['results']:
        check(len(row) == 3 and row[1] in ('正常', '疑似抄袭', '无图片'),
              f'{row[0]}: 判定 "{row[1]}"')
        if result['total'] > 0:
            print(f'    {row[0]}: {row[1]} — {row[2]}')


# ════════════════════════════════════════════
section('Test 4: run_pipeline 完整流程')
# ════════════════════════════════════════════

with tempfile.TemporaryDirectory() as td:
    logs = []
    result = run_pipeline(ZIP_PATH, td, log_func=lambda msg: logs.append(msg))

    check_eq(result['total'], 10, '流水线结果 total = 10')
    check_eq(result['plag_count'], 6, '流水线结果 plag_count = 6')
    check_eq(result['none_count'], 1, '流水线结果 none_count = 1（陈十十）')
    check('results' in result, '结果包含 results')
    check('excel_buf' in result, '结果包含 excel_buf')
    check(len(logs) >= 4, f'日志回调被调用 >= 4 次（实际 {len(logs)} 次）')

    for l in logs:
        print(f'    {l}')

    # Excel 可读且内容正确
    result['excel_buf'].seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(result['excel_buf'])
    check('统计概览' in wb.sheetnames, 'Excel 包含统计概览表')
    check('详细名单' in wb.sheetnames, 'Excel 包含详细名单表')
    ws = wb['详细名单']
    rows = list(ws.iter_rows(values_only=True))
    check(len(rows) - 1 == result['total'],
          f'详细名单行数匹配 ({len(rows)-1} == {result["total"]})')
    wb.close()


# ════════════════════════════════════════════
section('Test 5: extract_features — .docx 图片提取')
# ════════════════════════════════════════════

with tempfile.TemporaryDirectory() as td:
    raw = os.path.join(td, "raw")
    target = os.path.join(td, "organized")
    unzip_process(ZIP_PATH, raw)
    organize_process(raw, target)

    docx_files = sorted([f for f in os.listdir(target) if f.endswith('.docx')])
    check_eq(len(docx_files), 10, 'organized 目录含 10 个 .docx 文件')
    for f in docx_files[:5]:
        feats = extract_features(os.path.join(target, f))
        print(f'    {f}: {len(feats)} 张图片')
        for md5, size in feats[:3]:
            print(f'      MD5:{md5[:12]}... size:{size}')
    # 9 份有图片，1 份无图片（陈十十占位）
    with_images = sum(1 for f in docx_files
                      if extract_features(os.path.join(target, f)))
    without_images = sum(1 for f in docx_files
                         if not extract_features(os.path.join(target, f)))
    check_eq(with_images, 9, '9 份 .docx 提取到图片')
    check_eq(without_images, 1, '1 份无图片（陈十十空白占位）')


# ════════════════════════════════════════════
section(f'SUMMARY\n  report_checker e2e: {PASS} 通过, {FAIL} 失败')
if FAIL:
    sys.exit(1)
